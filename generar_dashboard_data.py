#!/usr/bin/env python3
"""
generar_dashboard_data.py
=========================
Genera data_dashboard.js con los mismos datos y filtros que usan los reportes.

FILTROS REPLICADOS (idénticos a los builders):
  - General:      total_benef de gran_total; apoyos de total_apoyos_excel;
                  programas con >= 10 beneficiarios; ap_clean excluye INST_NAMES_UP
  - Municipal:    desglose con total > 0; municipios reales (no especiales)
  - Institucional: apoyos de apoyos_g3; programas con >= 2 apoyos en prog_apoyos

Uso:
  python3 generar_dashboard_data.py <excel_path>         → genera data_dashboard.js
  python3 generar_dashboard_data.py <excel_path> --json  → imprime JSON puro (debug)

El HTML del dashboard importa data_dashboard.js para leer window.DASHBOARD_DATA.
"""

import sys, json, subprocess
from pathlib import Path

# ── Diccionario CONEVAL/INEGI estático por municipio ─────────────────────────
# Clave INEGI de 5 dígitos + indicadores CONEVAL 2020 (no cambian con el Excel)
# Indexado por nombre normalizado sin acentos en mayúsculas.
import unicodedata as _ud
def _norm_mun(s):
    s = s.upper().strip()
    return ''.join(c for c in _ud.normalize('NFD', s) if _ud.category(c) != 'Mn')

_INEGI_STATIC = {"CHIHUAHUA":{"clave":"08020","grado_dom":"Alto","pct_muy_alto":0,"pct_alto":60.0,"pct_medio":20.0,"pct_bajo":20.0,"pct_muy_bajo":0,"ageb_count":5,"avg_sin_internet":83.3,"avg_hacinamiento":12.2,"avg_analfabeta":6.3,"avg_piso_tierra":10.3,"avg_sin_salud":11.7,"idx_vulnerabilidad":57.7,"idx_dependencia":94.6,"idx_urgencia":80.3},"JUAREZ":{"clave":"08038","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":14.3,"pct_medio":85.7,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":7,"avg_sin_internet":74.6,"avg_hacinamiento":5.8,"avg_analfabeta":4.8,"avg_piso_tierra":0.9,"avg_sin_salud":5.0,"idx_vulnerabilidad":39.6,"idx_dependencia":61.0,"idx_urgencia":67.4},"CUAUHTEMOC":{"clave":"08018","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":1,"avg_sin_internet":81.1,"avg_hacinamiento":2.7,"avg_analfabeta":2.6,"avg_piso_tierra":0,"avg_sin_salud":13.5,"idx_vulnerabilidad":38.2,"idx_dependencia":66.4,"idx_urgencia":64.7},"HIDALGO DEL PARRAL":{"clave":"08033","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":1,"avg_sin_internet":61.5,"avg_hacinamiento":3.3,"avg_analfabeta":2.9,"avg_piso_tierra":0,"avg_sin_salud":2.9,"idx_vulnerabilidad":33.8,"idx_dependencia":56.7,"idx_urgencia":64.1},"DELICIAS":{"clave":"08022","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":5,"avg_sin_internet":49.5,"avg_hacinamiento":2.5,"avg_analfabeta":3.6,"avg_piso_tierra":0,"avg_sin_salud":10.3,"idx_vulnerabilidad":32.3,"idx_dependencia":48.4,"idx_urgencia":66.6},"CAMARGO":{"clave":"08012","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":40.0,"pct_medio":60.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":5,"avg_sin_internet":72.0,"avg_hacinamiento":7.3,"avg_analfabeta":6.8,"avg_piso_tierra":1.6,"avg_sin_salud":10.9,"idx_vulnerabilidad":42.2,"idx_dependencia":65.8,"idx_urgencia":59.6},"JIMENEZ":{"clave":"08037","grado_dom":"Medio","pct_muy_alto":0.8,"pct_alto":6.8,"pct_medio":40.4,"pct_bajo":37.3,"pct_muy_bajo":14.7,"ageb_count":633,"avg_sin_internet":42.3,"avg_hacinamiento":4.0,"avg_analfabeta":1.8,"avg_piso_tierra":2.2,"avg_sin_salud":21.7,"idx_vulnerabilidad":32.4,"idx_dependencia":67.4,"idx_urgencia":57.5},"GUERRERO":{"clave":"08032","grado_dom":"Bajo","pct_muy_alto":3.7,"pct_alto":2.5,"pct_medio":27.2,"pct_bajo":49.4,"pct_muy_bajo":17.3,"ageb_count":81,"avg_sin_internet":40.1,"avg_hacinamiento":2.6,"avg_analfabeta":1.9,"avg_piso_tierra":1.1,"avg_sin_salud":12.1,"idx_vulnerabilidad":22.9,"idx_dependencia":71.2,"idx_urgencia":44.2},"OJINAGA":{"clave":"08053","grado_dom":"Medio","pct_muy_alto":8.0,"pct_alto":0,"pct_medio":88.0,"pct_bajo":4.0,"pct_muy_bajo":0,"ageb_count":25,"avg_sin_internet":67.9,"avg_hacinamiento":6.1,"avg_analfabeta":7.4,"avg_piso_tierra":1.2,"avg_sin_salud":15.3,"idx_vulnerabilidad":41.5,"idx_dependencia":54.6,"idx_urgencia":56.6},"ALDAMA":{"clave":"08002","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":65.0,"pct_bajo":27.5,"pct_muy_bajo":7.5,"ageb_count":40,"avg_sin_internet":60.0,"avg_hacinamiento":3.2,"avg_analfabeta":2.3,"avg_piso_tierra":1.8,"avg_sin_salud":9.5,"idx_vulnerabilidad":34.5,"idx_dependencia":80.0,"idx_urgencia":58.2},"GUACHOCHI":{"clave":"08028","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":8,"avg_sin_internet":75.1,"avg_hacinamiento":10.0,"avg_analfabeta":7.4,"avg_piso_tierra":0.7,"avg_sin_salud":15.2,"idx_vulnerabilidad":44.4,"idx_dependencia":56.3,"idx_urgencia":64.4},"GUADALUPE Y CALVO":{"clave":"08030","grado_dom":"Medio","pct_muy_alto":14.3,"pct_alto":14.3,"pct_medio":42.9,"pct_bajo":14.3,"pct_muy_bajo":14.3,"ageb_count":7,"avg_sin_internet":80.9,"avg_hacinamiento":6.3,"avg_analfabeta":5.1,"avg_piso_tierra":2.4,"avg_sin_salud":20.0,"idx_vulnerabilidad":43.5,"idx_dependencia":70.0,"idx_urgencia":63.8},"BOCOYNA":{"clave":"08009","grado_dom":"Alto","pct_muy_alto":20.5,"pct_alto":41.0,"pct_medio":30.8,"pct_bajo":2.6,"pct_muy_bajo":5.1,"ageb_count":39,"avg_sin_internet":86.2,"avg_hacinamiento":17.2,"avg_analfabeta":6.5,"avg_piso_tierra":5.5,"avg_sin_salud":11.8,"idx_vulnerabilidad":58.3,"idx_dependencia":75.0,"idx_urgencia":72.0},"NUEVO CASAS GRANDES":{"clave":"08051","grado_dom":"Alto","pct_muy_alto":0,"pct_alto":100.0,"pct_medio":0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":1,"avg_sin_internet":97.7,"avg_hacinamiento":8.4,"avg_analfabeta":8.8,"avg_piso_tierra":3.8,"avg_sin_salud":33.3,"idx_vulnerabilidad":59.7,"idx_dependencia":61.6,"idx_urgencia":81.2},"MEOQUI":{"clave":"08046","grado_dom":"Muy Alto","pct_muy_alto":50.0,"pct_alto":50.0,"pct_medio":0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":2,"avg_sin_internet":95.5,"avg_hacinamiento":12.4,"avg_analfabeta":12.5,"avg_piso_tierra":14.1,"avg_sin_salud":11.7,"idx_vulnerabilidad":73.8,"idx_dependencia":69.1,"idx_urgencia":94.8},"MADERA":{"clave":"08041","grado_dom":"Muy Alto","pct_muy_alto":50.0,"pct_alto":50.0,"pct_medio":0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":4,"avg_sin_internet":99.3,"avg_hacinamiento":25.7,"avg_analfabeta":9.2,"avg_piso_tierra":2.7,"avg_sin_salud":3.4,"idx_vulnerabilidad":71.5,"idx_dependencia":63.0,"idx_urgencia":91.5},"ASCENSION":{"clave":"08005","grado_dom":"Medio","pct_muy_alto":4.1,"pct_alto":22.4,"pct_medio":63.3,"pct_bajo":2.0,"pct_muy_bajo":8.2,"ageb_count":49,"avg_sin_internet":67.6,"avg_hacinamiento":11.3,"avg_analfabeta":7.9,"avg_piso_tierra":2.8,"avg_sin_salud":36.4,"idx_vulnerabilidad":47.1,"idx_dependencia":58.7,"idx_urgencia":62.9},"NAMIQUIPA":{"clave":"08049","grado_dom":"Medio","pct_muy_alto":16.7,"pct_alto":16.7,"pct_medio":66.7,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":6,"avg_sin_internet":83.9,"avg_hacinamiento":5.6,"avg_analfabeta":7.2,"avg_piso_tierra":2.7,"avg_sin_salud":7.3,"idx_vulnerabilidad":44.3,"idx_dependencia":91.1,"idx_urgencia":61.2},"BUENAVENTURA":{"clave":"08011","grado_dom":"Bajo","pct_muy_alto":0,"pct_alto":4.7,"pct_medio":37.2,"pct_bajo":44.2,"pct_muy_bajo":14.0,"ageb_count":43,"avg_sin_internet":50.3,"avg_hacinamiento":2.2,"avg_analfabeta":2.5,"avg_piso_tierra":1.0,"avg_sin_salud":16.6,"idx_vulnerabilidad":25.6,"idx_dependencia":69.4,"idx_urgencia":48.2},"SAUCILLO":{"clave":"08063","grado_dom":"Medio","pct_muy_alto":20.0,"pct_alto":20.0,"pct_medio":60.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":5,"avg_sin_internet":88.2,"avg_hacinamiento":12.0,"avg_analfabeta":7.1,"avg_piso_tierra":0.4,"avg_sin_salud":9.5,"idx_vulnerabilidad":47.0,"idx_dependencia":65.9,"idx_urgencia":64.0},"URIQUE":{"clave":"08066","grado_dom":"Alto","pct_muy_alto":0,"pct_alto":66.7,"pct_medio":33.3,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":3,"avg_sin_internet":99.4,"avg_hacinamiento":10.9,"avg_analfabeta":6.5,"avg_piso_tierra":8.6,"avg_sin_salud":4.3,"idx_vulnerabilidad":58.9,"idx_dependencia":68.3,"idx_urgencia":74.8},"BALLEZA":{"clave":"08007","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":23.5,"pct_medio":47.1,"pct_bajo":17.6,"pct_muy_bajo":11.8,"ageb_count":17,"avg_sin_internet":60.4,"avg_hacinamiento":6.3,"avg_analfabeta":7.8,"avg_piso_tierra":1.5,"avg_sin_salud":10.2,"idx_vulnerabilidad":40.0,"idx_dependencia":73.6,"idx_urgencia":61.0},"AHUMADA":{"clave":"08001","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":4.5,"pct_medio":68.2,"pct_bajo":27.3,"pct_muy_bajo":0,"ageb_count":22,"avg_sin_internet":50.0,"avg_hacinamiento":2.9,"avg_analfabeta":2.8,"avg_piso_tierra":1.2,"avg_sin_salud":14.8,"idx_vulnerabilidad":33.0,"idx_dependencia":68.7,"idx_urgencia":61.1},"BATOPILAS DE MANUEL GOMEZ MORIN":{"clave":"08008","grado_dom":"Alto","pct_muy_alto":42.9,"pct_alto":57.1,"pct_medio":0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":7,"avg_sin_internet":95.7,"avg_hacinamiento":15.1,"avg_analfabeta":14.5,"avg_piso_tierra":34.9,"avg_sin_salud":13.0,"idx_vulnerabilidad":72.6,"idx_dependencia":76.5,"idx_urgencia":72.8},"VALLE DE ZARAGOZA":{"clave":"08068","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":0,"avg_sin_internet":0,"avg_hacinamiento":0,"avg_analfabeta":0,"avg_piso_tierra":0,"avg_sin_salud":0,"idx_vulnerabilidad":17.5,"idx_dependencia":94.2,"idx_urgencia":42.6},"CASAS GRANDES":{"clave":"08014","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":6,"avg_sin_internet":75.7,"avg_hacinamiento":3.0,"avg_analfabeta":3.5,"avg_piso_tierra":0,"avg_sin_salud":15.0,"idx_vulnerabilidad":38.1,"idx_dependencia":88.0,"idx_urgencia":60.3},"GUAZAPARES":{"clave":"08031","grado_dom":"Medio","pct_muy_alto":7.4,"pct_alto":22.2,"pct_medio":46.3,"pct_bajo":20.4,"pct_muy_bajo":3.7,"ageb_count":54,"avg_sin_internet":71.5,"avg_hacinamiento":6.5,"avg_analfabeta":4.7,"avg_piso_tierra":6.4,"avg_sin_salud":14.4,"idx_vulnerabilidad":42.9,"idx_dependencia":72.9,"idx_urgencia":57.9},"GOMEZ FARIAS":{"clave":"08026","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":16.7,"pct_medio":83.3,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":6,"avg_sin_internet":46.5,"avg_hacinamiento":2.8,"avg_analfabeta":5.4,"avg_piso_tierra":1.5,"avg_sin_salud":4.6,"idx_vulnerabilidad":33.3,"idx_dependencia":71.5,"idx_urgencia":53.7},"ROSALES":{"clave":"08056","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":2,"avg_sin_internet":55.5,"avg_hacinamiento":0,"avg_analfabeta":7.0,"avg_piso_tierra":0,"avg_sin_salud":15.7,"idx_vulnerabilidad":35.4,"idx_dependencia":70.8,"idx_urgencia":64.6},"CARICHI":{"clave":"08013","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":64.7,"pct_bajo":29.4,"pct_muy_bajo":5.9,"ageb_count":17,"avg_sin_internet":58.4,"avg_hacinamiento":2.9,"avg_analfabeta":3.0,"avg_piso_tierra":2.5,"avg_sin_salud":21.8,"idx_vulnerabilidad":36.2,"idx_dependencia":87.2,"idx_urgencia":58.6},"SANTA BARBARA":{"clave":"08061","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":2,"avg_sin_internet":91.4,"avg_hacinamiento":17.8,"avg_analfabeta":4.3,"avg_piso_tierra":0.8,"avg_sin_salud":1.2,"idx_vulnerabilidad":47.5,"idx_dependencia":70.3,"idx_urgencia":63.2},"JANOS":{"clave":"08036","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":7.4,"pct_medio":51.9,"pct_bajo":37.0,"pct_muy_bajo":3.7,"ageb_count":27,"avg_sin_internet":56.9,"avg_hacinamiento":2.7,"avg_analfabeta":2.8,"avg_piso_tierra":1.2,"avg_sin_salud":12.6,"idx_vulnerabilidad":34.1,"idx_dependencia":72.8,"idx_urgencia":62.6},"LOPEZ":{"clave":"08040","grado_dom":"Medio","pct_muy_alto":10.4,"pct_alto":19.4,"pct_medio":56.7,"pct_bajo":9.0,"pct_muy_bajo":4.5,"ageb_count":67,"avg_sin_internet":83.4,"avg_hacinamiento":6.8,"avg_analfabeta":6.2,"avg_piso_tierra":6.8,"avg_sin_salud":13.0,"idx_vulnerabilidad":46.6,"idx_dependencia":63.0,"idx_urgencia":50.4},"BACHINIVA":{"clave":"08006","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":7.7,"pct_medio":92.3,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":13,"avg_sin_internet":72.5,"avg_hacinamiento":6.7,"avg_analfabeta":3.8,"avg_piso_tierra":2.7,"avg_sin_salud":13.4,"idx_vulnerabilidad":40.6,"idx_dependencia":85.3,"idx_urgencia":56.1},"IGNACIO ZARAGOZA":{"clave":"08035","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":10.5,"pct_medio":73.7,"pct_bajo":15.8,"pct_muy_bajo":0,"ageb_count":19,"avg_sin_internet":43.7,"avg_hacinamiento":11.2,"avg_analfabeta":7.6,"avg_piso_tierra":6.4,"avg_sin_salud":38.8,"idx_vulnerabilidad":44.1,"idx_dependencia":97.1,"idx_urgencia":54.3},"SANTA ISABEL":{"clave":"08025","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":9.1,"pct_medio":90.9,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":11,"avg_sin_internet":86.3,"avg_hacinamiento":2.6,"avg_analfabeta":6.4,"avg_piso_tierra":0.6,"avg_sin_salud":10.6,"idx_vulnerabilidad":42.1,"idx_dependencia":79.9,"idx_urgencia":50.6},"MATAMOROS":{"clave":"08045","grado_dom":"Medio","pct_muy_alto":2.4,"pct_alto":14.3,"pct_medio":42.9,"pct_bajo":31.0,"pct_muy_bajo":9.5,"ageb_count":42,"avg_sin_internet":56.0,"avg_hacinamiento":6.3,"avg_analfabeta":3.8,"avg_piso_tierra":2.5,"avg_sin_salud":15.0,"idx_vulnerabilidad":37.1,"idx_dependencia":92.3,"idx_urgencia":53.3},"PRAXEDIS G. GUERRERO":{"clave":"08054","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":2,"avg_sin_internet":84.9,"avg_hacinamiento":1.6,"avg_analfabeta":1.1,"avg_piso_tierra":0.9,"avg_sin_salud":7.2,"idx_vulnerabilidad":37.2,"idx_dependencia":93.1,"idx_urgencia":55.0},"OCAMPO":{"clave":"08052","grado_dom":"Medio","pct_muy_alto":2.1,"pct_alto":0,"pct_medio":79.2,"pct_bajo":18.8,"pct_muy_bajo":0,"ageb_count":48,"avg_sin_internet":58.3,"avg_hacinamiento":3.6,"avg_analfabeta":2.6,"avg_piso_tierra":1.9,"avg_sin_salud":16.8,"idx_vulnerabilidad":35.4,"idx_dependencia":77.1,"idx_urgencia":62.0},"LA CRUZ":{"clave":"08017","grado_dom":"Medio","pct_muy_alto":2.0,"pct_alto":2.7,"pct_medio":37.6,"pct_bajo":36.2,"pct_muy_bajo":21.5,"ageb_count":149,"avg_sin_internet":46.3,"avg_hacinamiento":3.9,"avg_analfabeta":2.4,"avg_piso_tierra":0.6,"avg_sin_salud":12.9,"idx_vulnerabilidad":31.9,"idx_dependencia":81.2,"idx_urgencia":50.9},"URUACHI":{"clave":"08067","grado_dom":"Medio","pct_muy_alto":6.2,"pct_alto":0,"pct_medio":87.5,"pct_bajo":6.2,"pct_muy_bajo":0,"ageb_count":16,"avg_sin_internet":69.6,"avg_hacinamiento":1.7,"avg_analfabeta":2.2,"avg_piso_tierra":0.8,"avg_sin_salud":14.6,"idx_vulnerabilidad":35.7,"idx_dependencia":73.7,"idx_urgencia":60.8},"MORIS":{"clave":"08048","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":5.7,"pct_medio":85.7,"pct_bajo":2.9,"pct_muy_bajo":5.7,"ageb_count":35,"avg_sin_internet":79.1,"avg_hacinamiento":4.6,"avg_analfabeta":3.6,"avg_piso_tierra":2.6,"avg_sin_salud":18.8,"idx_vulnerabilidad":41.3,"idx_dependencia":85.9,"idx_urgencia":55.8},"JULIMES":{"clave":"08039","grado_dom":"Bajo","pct_muy_alto":0,"pct_alto":0,"pct_medio":0,"pct_bajo":100.0,"pct_muy_bajo":0,"ageb_count":3,"avg_sin_internet":31.5,"avg_hacinamiento":2.2,"avg_analfabeta":2.1,"avg_piso_tierra":1.3,"avg_sin_salud":4.4,"idx_vulnerabilidad":20.5,"idx_dependencia":71.4,"idx_urgencia":44.4},"TEMOSACHIC":{"clave":"08064","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":5,"avg_sin_internet":45.3,"avg_hacinamiento":5.3,"avg_analfabeta":9.0,"avg_piso_tierra":3.6,"avg_sin_salud":6.4,"idx_vulnerabilidad":38.1,"idx_dependencia":78.2,"idx_urgencia":58.7},"CHINIPAS":{"clave":"08021","grado_dom":"Bajo","pct_muy_alto":1.9,"pct_alto":2.8,"pct_medio":25.9,"pct_bajo":50.0,"pct_muy_bajo":19.4,"ageb_count":108,"avg_sin_internet":40.0,"avg_hacinamiento":3.0,"avg_analfabeta":2.1,"avg_piso_tierra":2.0,"avg_sin_salud":13.6,"idx_vulnerabilidad":23.8,"idx_dependencia":57.5,"idx_urgencia":45.4},"SATEVO":{"clave":"08062","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":4.0,"pct_medio":48.0,"pct_bajo":36.0,"pct_muy_bajo":12.0,"ageb_count":25,"avg_sin_internet":55.6,"avg_hacinamiento":2.4,"avg_analfabeta":2.9,"avg_piso_tierra":0.8,"avg_sin_salud":12.6,"idx_vulnerabilidad":33.5,"idx_dependencia":125.4,"idx_urgencia":56.6},"SAN FRANCISCO DEL ORO":{"clave":"08060","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":5.3,"pct_medio":52.6,"pct_bajo":26.3,"pct_muy_bajo":15.8,"ageb_count":19,"avg_sin_internet":55.8,"avg_hacinamiento":3.3,"avg_analfabeta":3.6,"avg_piso_tierra":1.3,"avg_sin_salud":7.4,"idx_vulnerabilidad":34.2,"idx_dependencia":81.6,"idx_urgencia":60.3},"CORONADO":{"clave":"08015","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":9,"avg_sin_internet":91.9,"avg_hacinamiento":1.1,"avg_analfabeta":3.9,"avg_piso_tierra":1.1,"avg_sin_salud":16.7,"idx_vulnerabilidad":41.5,"idx_dependencia":83.6,"idx_urgencia":46.4},"SAN FRANCISCO DE CONCHOS":{"clave":"08059","grado_dom":"Medio","pct_muy_alto":7.7,"pct_alto":0,"pct_medio":53.8,"pct_bajo":30.8,"pct_muy_bajo":7.7,"ageb_count":13,"avg_sin_internet":55.3,"avg_hacinamiento":1.1,"avg_analfabeta":3.3,"avg_piso_tierra":2.2,"avg_sin_salud":3.3,"idx_vulnerabilidad":33.0,"idx_dependencia":71.2,"idx_urgencia":54.0},"AQUILES SERDAN":{"clave":"08004","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":12,"avg_sin_internet":68.2,"avg_hacinamiento":1.6,"avg_analfabeta":1.3,"avg_piso_tierra":1.1,"avg_sin_salud":10.4,"idx_vulnerabilidad":34.4,"idx_dependencia":54.2,"idx_urgencia":68.6},"GUADALUPE":{"clave":"08029","grado_dom":"Alto","pct_muy_alto":37.1,"pct_alto":40.0,"pct_medio":20.0,"pct_bajo":2.9,"pct_muy_bajo":0,"ageb_count":35,"avg_sin_internet":90.4,"avg_hacinamiento":15.8,"avg_analfabeta":14.2,"avg_piso_tierra":15.5,"avg_sin_salud":5.6,"idx_vulnerabilidad":68.7,"idx_dependencia":71.3,"idx_urgencia":72.5},"GALEANA":{"clave":"08024","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":11.1,"pct_medio":44.4,"pct_bajo":33.3,"pct_muy_bajo":11.1,"ageb_count":9,"avg_sin_internet":49.1,"avg_hacinamiento":1.8,"avg_analfabeta":3.1,"avg_piso_tierra":0.6,"avg_sin_salud":11.9,"idx_vulnerabilidad":31.9,"idx_dependencia":67.4,"idx_urgencia":63.9},"MATACHI":{"clave":"08044","grado_dom":"Medio","pct_muy_alto":12.5,"pct_alto":0,"pct_medio":75.0,"pct_bajo":6.2,"pct_muy_bajo":6.2,"ageb_count":16,"avg_sin_internet":80.6,"avg_hacinamiento":2.2,"avg_analfabeta":8.8,"avg_piso_tierra":11.2,"avg_sin_salud":8.7,"idx_vulnerabilidad":47.7,"idx_dependencia":89.9,"idx_urgencia":54.4},"EL TULE":{"clave":"08065","grado_dom":"Alto","pct_muy_alto":0,"pct_alto":100.0,"pct_medio":0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":2,"avg_sin_internet":99.6,"avg_hacinamiento":12.9,"avg_analfabeta":6.7,"avg_piso_tierra":12.9,"avg_sin_salud":9.8,"idx_vulnerabilidad":62.7,"idx_dependencia":98.4,"idx_urgencia":54.9},"NONOAVA":{"clave":"08050","grado_dom":"Bajo","pct_muy_alto":0,"pct_alto":5.5,"pct_medio":40.0,"pct_bajo":47.3,"pct_muy_bajo":7.3,"ageb_count":55,"avg_sin_internet":46.4,"avg_hacinamiento":1.5,"avg_analfabeta":1.8,"avg_piso_tierra":0.5,"avg_sin_salud":19.8,"idx_vulnerabilidad":24.0,"idx_dependencia":82.7,"idx_urgencia":42.8},"SAN FRANCISCO DE BORJA":{"clave":"08058","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":2,"avg_sin_internet":64.7,"avg_hacinamiento":2.0,"avg_analfabeta":1.6,"avg_piso_tierra":0,"avg_sin_salud":5.0,"idx_vulnerabilidad":33.0,"idx_dependencia":92.6,"idx_urgencia":54.0},"MAGUARICHI":{"clave":"08042","grado_dom":"Medio","pct_muy_alto":20.0,"pct_alto":0,"pct_medio":60.0,"pct_bajo":20.0,"pct_muy_bajo":0,"ageb_count":5,"avg_sin_internet":88.2,"avg_hacinamiento":0,"avg_analfabeta":10.4,"avg_piso_tierra":0.5,"avg_sin_salud":2.4,"idx_vulnerabilidad":43.4,"idx_dependencia":91.2,"idx_urgencia":42.0},"GRAN MORELOS":{"clave":"08027","grado_dom":"Medio","pct_muy_alto":9.4,"pct_alto":34.4,"pct_medio":46.9,"pct_bajo":9.4,"pct_muy_bajo":0,"ageb_count":32,"avg_sin_internet":78.7,"avg_hacinamiento":10.2,"avg_analfabeta":5.0,"avg_piso_tierra":2.4,"avg_sin_salud":12.4,"idx_vulnerabilidad":44.0,"idx_dependencia":65.7,"idx_urgencia":56.7},"COYAME DEL SOTOL":{"clave":"08016","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":12.5,"pct_medio":87.5,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":8,"avg_sin_internet":70.2,"avg_hacinamiento":5.9,"avg_analfabeta":6.1,"avg_piso_tierra":2.0,"avg_sin_salud":10.5,"idx_vulnerabilidad":40.8,"idx_dependencia":98.7,"idx_urgencia":46.4},"CUSIHUIRIACHI":{"clave":"08019","grado_dom":"Bajo","pct_muy_alto":1.1,"pct_alto":4.6,"pct_medio":23.4,"pct_bajo":38.2,"pct_muy_bajo":32.8,"ageb_count":659,"avg_sin_internet":35.8,"avg_hacinamiento":3.4,"avg_analfabeta":1.7,"avg_piso_tierra":1.9,"avg_sin_salud":12.1,"idx_vulnerabilidad":22.6,"idx_dependencia":69.1,"idx_urgencia":50.3},"MANUEL BENAVIDES":{"clave":"08043","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":100.0,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":6,"avg_sin_internet":84.8,"avg_hacinamiento":5.5,"avg_analfabeta":3.4,"avg_piso_tierra":0.6,"avg_sin_salud":15.4,"idx_vulnerabilidad":41.3,"idx_dependencia":67.0,"idx_urgencia":45.9},"HUEJOTITAN":{"clave":"08034","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":8.7,"pct_medio":78.3,"pct_bajo":8.7,"pct_muy_bajo":4.3,"ageb_count":23,"avg_sin_internet":58.1,"avg_hacinamiento":6.8,"avg_analfabeta":4.7,"avg_piso_tierra":3.2,"avg_sin_salud":11.6,"idx_vulnerabilidad":38.5,"idx_dependencia":84.6,"idx_urgencia":37.8},"DR. BELISARIO DOMINGUEZ":{"clave":"08023","grado_dom":"Medio","pct_muy_alto":8.3,"pct_alto":8.3,"pct_medio":83.3,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":12,"avg_sin_internet":81.1,"avg_hacinamiento":8.4,"avg_analfabeta":8.1,"avg_piso_tierra":12.5,"avg_sin_salud":35.9,"idx_vulnerabilidad":53.4,"idx_dependencia":163.9,"idx_urgencia":59.9},"RIVA PALACIO":{"clave":"08055","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":0,"pct_medio":84.6,"pct_bajo":15.4,"pct_muy_bajo":0,"ageb_count":13,"avg_sin_internet":49.0,"avg_hacinamiento":3.9,"avg_analfabeta":2.8,"avg_piso_tierra":0.7,"avg_sin_salud":12.4,"idx_vulnerabilidad":32.7,"idx_dependencia":75.4,"idx_urgencia":67.0},"ALLENDE":{"clave":"08003","grado_dom":"Bajo","pct_muy_alto":7.4,"pct_alto":11.1,"pct_medio":25.9,"pct_bajo":37.0,"pct_muy_bajo":18.5,"ageb_count":27,"avg_sin_internet":60.9,"avg_hacinamiento":6.8,"avg_analfabeta":6.4,"avg_piso_tierra":4.4,"avg_sin_salud":20.7,"idx_vulnerabilidad":34.8,"idx_dependencia":62.3,"idx_urgencia":52.8},"MORELOS":{"clave":"08047","grado_dom":"Alto","pct_muy_alto":11.1,"pct_alto":66.7,"pct_medio":22.2,"pct_bajo":0,"pct_muy_bajo":0,"ageb_count":9,"avg_sin_internet":98.6,"avg_hacinamiento":11.5,"avg_analfabeta":7.8,"avg_piso_tierra":5.5,"avg_sin_salud":16.1,"idx_vulnerabilidad":59.6,"idx_dependencia":60.0,"idx_urgencia":82.7},"ROSARIO":{"clave":"08057","grado_dom":"Medio","pct_muy_alto":0,"pct_alto":20.0,"pct_medio":60.0,"pct_bajo":20.0,"pct_muy_bajo":0,"ageb_count":5,"avg_sin_internet":56.6,"avg_hacinamiento":2.5,"avg_analfabeta":5.4,"avg_piso_tierra":0.8,"avg_sin_salud":4.3,"idx_vulnerabilidad":34.8,"idx_dependencia":88.2,"idx_urgencia":63.0}}


SCRIPT_DIR = Path(__file__).parent
READER     = SCRIPT_DIR / 'read_excel_padron.py'

# ── Constantes (mismas que motor_reporte_padron.py) ──────────────────────────
POB_ESTATAL    = 4_043_130
POB_VULNERABLE = 1_792_324

# Instituciones/nombres excluidos del listado de tipos de apoyo (mismo filtro que motor_reporte_padron.py)
INST_NAMES_UP = {
    'CECYTECH','COESPO','COESVI','DIF','ICHD','ICHDII','ICHIJUV','ICHIMUJ',
    'RURAL','SALUD','SDBYBC','SDHyBC','SDHYBC','SEECH','SEYD','SEyD',
    'SPyCI','SPYCI','TRABAJO','TURISMO','CULTURA','MEDICHIHUAHUA',
    'DESARROLLO HUMANO','NO IDENTIFICADO',
}

def sf(v):
    try: return float(v or 0)
    except: return 0.0

def pct(part, total):
    if not total: return 0.0
    return round(float(part) / float(total) * 100, 1)

# ── Leer datos crudos del Excel ───────────────────────────────────────────────
def leer_excel(excel_path):
    result = subprocess.run(
        ['python3', str(READER), str(excel_path)],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print('ERROR al leer el Excel:', result.stderr, file=sys.stderr)
        sys.exit(1)
    return json.loads(result.stdout)

# ── Construir payload del dashboard ──────────────────────────────────────────
def leer_grupos_vulnerables(excel_path):
    """Lee la hoja Grupos Vulnerables directamente con openpyxl.
    Retorna: {
        'grupos': [{'nombre': str, 'pob_vulnerable': int, 'atendidos': int}, ...],
        'mujeres': {'pob_vulnerable': int, 'atendidas': int},
        'hombres': {'pob_vulnerable': int, 'atendidos': int},
    }
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        if 'Grupos Vulnerables' not in wb.sheetnames:
            return {}
        ws = wb['Grupos Vulnerables']
        result = {'grupos': [], 'mujeres': {}, 'hombres': {}}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            nombre = str(row[0]).strip()
            try:
                pob_vul = int(row[1]) if row[1] not in (None, '') else 0
            except (ValueError, TypeError):
                pob_vul = 0
            try:
                pob_ate = int(row[2]) if row[2] not in (None, '') else 0
            except (ValueError, TypeError):
                pob_ate = 0
            nombre_l = nombre.lower()
            if 'muj' in nombre_l:
                result['mujeres'] = {'pob_vulnerable': pob_vul, 'atendidas': pob_ate}
            elif 'hom' in nombre_l:
                result['hombres'] = {'pob_vulnerable': pob_vul, 'atendidos': pob_ate}
            result['grupos'].append({
                'nombre': nombre,
                'pob_vulnerable': pob_vul,
                'atendidos': pob_ate,
            })
        return result
    except Exception as e:
        print(f'AVISO: No se pudo leer Grupos Vulnerables: {e}', file=sys.stderr)
        return {}


def leer_nutrichihuahua(excel_path):
    """Lee la hoja Nutrichihuahua directamente con openpyxl."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        if 'Nutrichihuahua' not in wb.sheetnames:
            return {}
        ws = wb['Nutrichihuahua']
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                result[str(row[0]).strip()] = int(row[1]) if row[1] else 0
        return result
    except Exception as e:
        print(f'AVISO: No se pudo leer Nutrichihuahua: {e}', file=sys.stderr)
        return {}


def build_dashboard_data(raw, excel_path=None):
    gt            = raw['gran_total']
    rangos        = raw['rangos_edad']            # col de rangos globales
    rangos_mh     = raw.get('rangos_mh_global', {})  # col S: desglose M/H
    instituciones = raw['instituciones']          # 5 instituciones principales
    municipios    = raw['municipios']             # lista completa
    apoyos        = raw['apoyos']                 # listado de tipos de apoyo
    loc           = raw.get('localizables', {})
    indicadores   = raw.get('indicadores', [])
    apoyos_g3     = raw.get('apoyos_g3', {})      # apoyos por institución (hoja 3)
    desglose_mun  = raw.get('desglose_municipal', {})  # {mun: [entries]}

    # ══ FILTROS REPORTE GENERAL ══════════════════════════════════════════════

    # Beneficiarios únicos — directo de gran_total (mismo que reporte)
    total_benef = int(sf(gt.get('total', 0)))
    total_m     = int(sf(gt.get('m', 0)))
    total_h     = int(sf(gt.get('h', 0)))
    total_sn    = int(sum(sf(v.get('sn', 0)) for v in instituciones.values()))

    # Apoyos — total directo del Excel (fila TOTAL de Apoyos Otorgados)
    total_apoyos = int(sf(raw.get('total_apoyos_excel', 0)))

    # Instituciones activas (>= 10 beneficiarios)
    inst_act = [k for k, v in instituciones.items() if sf(v.get('total', 0)) >= 10]
    total_inst = len(inst_act)
    total_prog = sum(len(v.get('programas', [])) for v in instituciones.values())

    # Tipos de apoyo limpios (excluye nombres de instituciones/placeholders)
    ap_clean = [a for a in apoyos
                if str(a.get('apoyo', '')).upper().strip() not in INST_NAMES_UP]

    # ── Rangos de edad (fuente: col S del Excel, con desglose M/H) ──────────
    edad_labels = [
        ('0 - 5 años',       '0-5'),
        ('6 - 11 años',      '6-11'),
        ('12 - 17 años',     '12-17'),
        ('18 - 29 años',     '18-29'),
        ('30 - 49 años',     '30-49'),
        ('50 - 64 años',     '50-64'),
        ('65 años o más',    '65+'),
        ('Sin dato de edad', 'sin_datos'),
    ]
    rangos_data = []
    for label, key in edad_labels:
        t_e  = int(sf(rangos.get(key, 0)))
        mh   = rangos_mh.get(key, {})
        m_e  = int(sf(mh.get('m', 0)))
        h_e  = int(sf(mh.get('h', 0)))
        sn_e = int(sf(mh.get('sn', max(0, t_e - m_e - h_e))))
        rangos_data.append({
            'label': label, 'key': key,
            'total': t_e, 'm': m_e, 'h': h_e, 'sn': sn_e,
            'pct_total': pct(t_e, total_benef),
        })

    # Grupos de edad simplificados (igual que reporte)
    ninos   = int(sf(rangos.get('0-5', 0)))   + int(sf(rangos.get('6-11', 0)))
    jovenes = int(sf(rangos.get('12-17', 0))) + int(sf(rangos.get('18-29', 0)))
    adultos = int(sf(rangos.get('30-49', 0))) + int(sf(rangos.get('50-64', 0)))
    mayores = int(sf(rangos.get('65+', 0)))

    # ── Localizables ─────────────────────────────────────────────────────────
    loc_total = int(loc.get('total', 0))
    loc_m     = int(loc.get('m', 0))
    loc_h     = int(loc.get('h', 0))
    loc_inst_raw = loc.get('por_institucion', [])  # [{nombre,m,h,total}]
    loc_rangos   = loc.get('rangos_edad', {})

    # Localizables por municipio — del objeto municipios (enriquecido por read_excel)
    loc_por_municipio = {}
    for mun in municipios:
        if not mun.get('especial'):
            nombre = mun['municipio']
            loc_por_municipio[nombre] = {
                'total': int(mun.get('total_localizables', 0)),
                'm':     int(mun.get('loc_m', 0)),
                'h':     int(mun.get('loc_h', 0)),
            }

    # Función de normalización (quitar tildes, mayúsculas)
    import unicodedata as _unicodedata_early
    def _nk(s): return _unicodedata_early.normalize('NFD',(s or '').upper()).encode('ascii','ignore').decode()

    # Abrir workbook y leer hoja S de una sola vez (programas e instituciones)
    hoja_s_rangos      = {}
    hoja_s_sin         = {}
    hoja_s_inst_rangos = {}
    hoja_s_inst_sin    = {}
    _wb_s = None
    if excel_path:
        try:
            import openpyxl as _opx_early
            _wb_s = _opx_early.load_workbook(str(excel_path), data_only=True, read_only=True)
        except Exception:
            _wb_s = None
    _rk2  = ['0-5','6-11','12-17','18-29','30-49','50-64','65+']
    if _wb_s and 'Unicos y Rango de Edad' in _wb_s.sheetnames:
        _ws_s_early  = _wb_s['Unicos y Rango de Edad']
        _rows_s_early = list(_ws_s_early.iter_rows(values_only=True))
        _COL_N_E, _COL_R0_E, _COL_SIN_E, _COL_TOT_E = 6, 7, 14, 15
        for _row_e in _rows_s_early:
            _nom_e = _row_e[_COL_N_E] if _COL_N_E < len(_row_e) else None
            if not isinstance(_nom_e, str) or not _nom_e.strip(): continue
            _nom_e = _nom_e.strip()
            if _nom_e in ('M','H','Sin datos','BENEFICIARIOS UNICOS POR RANGO DE EDAD','BENEFICIARIOS ÚNICOS '): continue
            _tot_e = _row_e[_COL_TOT_E] if _COL_TOT_E < len(_row_e) else None
            if not isinstance(_tot_e, (int, float)): continue
            _rangos_e = {k: (int(_row_e[_COL_R0_E+i]) if isinstance(_row_e[_COL_R0_E+i],(int,float)) and _row_e[_COL_R0_E+i] else 0)
                         for i,k in enumerate(_rk2) if (_COL_R0_E+i) < len(_row_e)}
            _sin_e = _row_e[_COL_SIN_E] if _COL_SIN_E < len(_row_e) else None
            _sin_e = int(_sin_e) if isinstance(_sin_e,(int,float)) and _sin_e else 0
            _key_e = _nk(_nom_e)
            hoja_s_rangos[_key_e]      = _rangos_e
            hoja_s_sin[_key_e]         = _sin_e
            hoja_s_inst_rangos[_key_e] = _rangos_e   # filas de inst también quedan aquí
            hoja_s_inst_sin[_key_e]    = _sin_e

    # ── Rangos por INSTITUCIÓN desde hoja S (filas de resumen) ──────────────
    if False and _wb_s and 'Unicos y Rango de Edad' in _wb_s.sheetnames:
        ws_inst = _wb_s['Unicos y Rango de Edad']
        rows_inst = list(ws_inst.iter_rows(values_only=True))
        for row_i in rows_inst:
            nombre_i = row_i[COL_N] if COL_N < len(row_i) else None
            if not isinstance(nombre_i, str) or not nombre_i.strip():
                continue
            nombre_i = nombre_i.strip()
            if nombre_i in ('M','H','Sin datos','BENEFICIARIOS UNICOS POR RANGO DE EDAD','BENEFICIARIOS ÚNICOS '):
                continue
            total_i = row_i[COL_TOT] if COL_TOT < len(row_i) else None
            if not isinstance(total_i, (int, float)):
                continue
            rangos_i = {}
            for i_r, k_r in enumerate(_rk2):
                v_r = row_i[COL_R0 + i_r] if (COL_R0 + i_r) < len(row_i) else None
                rangos_i[k_r] = int(v_r) if isinstance(v_r, (int, float)) and v_r else 0
            sin_i = row_i[COL_SIN] if COL_SIN < len(row_i) else None
            sin_i = int(sin_i) if isinstance(sin_i, (int, float)) and sin_i else 0
            key_i = _nk(nombre_i)
            hoja_s_inst_rangos[key_i] = rangos_i
            hoja_s_inst_sin[key_i]    = sin_i

    # ── Instituciones (para tab Instituciones) ────────────────────────────────
    instituciones_data = {}
    for inst_name, v in instituciones.items():
        tot = int(sf(v.get('total', 0)))
        if tot == 0:
            continue
        # Apoyos de esta institución desde apoyos_g3 (mismo que reporte general sec. 4)
        g3_inst         = apoyos_g3.get(inst_name, {})
        tot_apoyos_inst = int(sf(g3_inst.get('total', 0)))
        g3_progs        = g3_inst.get('programas', {})

        # Programas con apoyos >= 2 (filtro build_institucion.js línea 406)
        prog_apoyos_filtrado = {k: int(sf(v2)) for k, v2 in g3_progs.items()
                                if sf(v2) >= 2}

        programas = []
        for p in sorted(v.get('programas', []), key=lambda x: -sf(x.get('total', 0))):
            # Buscar apoyos con normalización tolerante a tildes
            def norm(s):
                import unicodedata
                return unicodedata.normalize('NFD', (s or '').upper()).encode('ascii', 'ignore').decode()
            ap_prog = prog_apoyos_filtrado.get(p['nombre'], 0)
            if not ap_prog:
                for k, val in prog_apoyos_filtrado.items():
                    if norm(k) == norm(p['nombre']):
                        ap_prog = val
                        break
            programas.append({
                'nombre':  p['nombre'],
                'total':   int(sf(p.get('total', 0))),
                'm':       int(sf(p.get('m', 0))),
                'h':       int(sf(p.get('h', 0))),
                'sn':      int(sf(p.get('sn', 0))),
                'apoyos':  ap_prog,
            })

        # Rangos: preferir hoja S (únicos) si disponible
        _inst_key = _nk(inst_name)
        rangos_hojas = hoja_s_inst_rangos.get(_inst_key, {})
        sin_hojas    = hoja_s_inst_sin.get(_inst_key, 0)
        rangos_raw   = v.get('rangos', {})
        use_rangos   = rangos_hojas if rangos_hojas else rangos_raw
        instituciones_data[inst_name] = {
            'total':       tot,
            'm':           int(sf(v.get('m', 0))),
            'h':           int(sf(v.get('h', 0))),
            'sn':          int(sf(v.get('sn', 0))),
            'apoyos':      tot_apoyos_inst,
            'programas':   programas,
            'rangos':      {k: int(sf(use_rangos.get(k, 0)))
                           for k in ['0-5','6-11','12-17','18-29','30-49','50-64','65+','sin_datos']},
            'sin_datos_edad': sin_hojas,
        }

    # ── Municipios (para tab Municipios) ──────────────────────────────────────
    # ── Inst por municipio (para capas G3 del mapa) ──────────────────────────
    # inst_subtotales viene directo de parse_sheet2: {inst_nombre: {m, h, total}}
    # Es la fuente más directa y confiable — ya está en cada objeto municipio
    # Construimos un lookup normalizado para usarlo en el loop de mun_reales
    inst_subtotales_por_mun = {}
    for m in municipios:
        nom_k = _norm_mun(m.get('municipio', ''))
        subtotales = m.get('inst_subtotales', {})
        if subtotales:
            inst_subtotales_por_mun[nom_k] = {
                ins: {'benef': int(sf(v.get('total', 0))),
                      'apoyos': int(sf(v.get('total', 0)))}
                for ins, v in subtotales.items()
                if sf(v.get('total', 0)) > 0
            }

    # Sólo municipios reales (no especiales), ordenados por volumen desc
    municipios_data = []
    mun_reales = sorted([m for m in municipios if not m.get('especial')],
                        key=lambda x: -x.get('total', 0))
    RANGO_KEYS = ['0-5','6-11','12-17','18-29','30-49','50-64','65+']

    for m in mun_reales:
        nom = m['municipio']
        # Rangos etarios del municipio (vienen del parser de la hoja Beneficiarios por Municipio)
        rg_raw = m.get('rangos', {})
        rg = {k: int(sf(rg_raw.get(k, 0))) for k in RANGO_KEYS}
        rg['sin_datos'] = int(sf(rg_raw.get('sin_datos', 0)))

        # Rango dominante (mayor cantidad, excluyendo sin_datos)
        rango_dom = max(RANGO_KEYS, key=lambda k: rg.get(k, 0)) if any(rg.get(k,0) for k in RANGO_KEYS) else None
        # Rango menor (con al menos 1 beneficiario, excluyendo sin_datos)
        rangos_con_dato = [k for k in RANGO_KEYS if rg.get(k, 0) > 0]
        rango_min = min(rangos_con_dato, key=lambda k: rg.get(k, 0)) if rangos_con_dato else None

        # Datos CONEVAL/INEGI estáticos para este municipio
        _st   = _INEGI_STATIC.get(_norm_mun(nom), {})
        _benef = int(m.get('total', 0))
        _apoyos = int(m.get('total_apoyos', 0))
        _pob  = int(m.get('poblacion', 0)) or _st.get('poblacion', 0)
        _loc  = int(m.get('total_localizables', 0))
        # Indicadores derivados del padrón
        _pct_65mas   = round(rg.get('65+', 0) / _benef * 100, 1)  if _benef else 0
        _pct_ninos   = round((rg.get('0-5', 0) + rg.get('6-11', 0)) / _benef * 100, 1) if _benef else 0
        _pct_jovenes = round(rg.get('18-29', 0) / _benef * 100, 1) if _benef else 0

        municipios_data.append({
            # ── Identificación ──
            'clave':          _st.get('clave', ''),
            'nombre':         nom,
            # ── Padrón ──
            'total':          _benef,
            'm':              int(m.get('m', 0)),
            'h':              int(m.get('h', 0)),
            'sn':             int(m.get('sn', 0)),
            'poblacion':      _pob,
            'total_apoyos':   _apoyos,
            'n_programas':    int(m.get('n_programas', 0)),
            'localizables':   _loc,
            'loc_m':          int(m.get('loc_m', 0)),
            'loc_h':          int(m.get('loc_h', 0)),
            'pct_loc':        round(_loc / _benef * 100, 1) if _benef else 0,
            'rangos':         rg,
            'rango_dom':      rango_dom,
            'rango_min':      rango_min,
            'pct_65mas':      _pct_65mas,
            'pct_ninos':      _pct_ninos,
            'pct_jovenes':    _pct_jovenes,
            # ── CONEVAL/INEGI estáticos ──
            'grado_dom':          _st.get('grado_dom', ''),
            'pct_muy_alto':       _st.get('pct_muy_alto', 0),
            'pct_alto':           _st.get('pct_alto', 0),
            'pct_medio':          _st.get('pct_medio', 0),
            'pct_bajo':           _st.get('pct_bajo', 0),
            'pct_muy_bajo':       _st.get('pct_muy_bajo', 0),
            'ageb_count':         _st.get('ageb_count', 0),
            'avg_sin_internet':   _st.get('avg_sin_internet', 0),
            'avg_hacinamiento':   _st.get('avg_hacinamiento', 0),
            'avg_analfabeta':     _st.get('avg_analfabeta', 0),
            'avg_piso_tierra':    _st.get('avg_piso_tierra', 0),
            'avg_sin_salud':      _st.get('avg_sin_salud', 0),
            'idx_vulnerabilidad': _st.get('idx_vulnerabilidad', 0),
            'idx_dependencia':    _st.get('idx_dependencia', 0),
            'idx_urgencia':       _st.get('idx_urgencia', 0),
            # ── Desglose por institución (para capas G3 del mapa) ──
            'inst':               inst_subtotales_por_mun.get(_norm_mun(nom), {}),
        })

    # Municipios especiales (Foráneo, No identificado)
    mun_especiales = []
    for m in [x for x in municipios if x.get('especial')]:
        mun_especiales.append({
            'nombre': m['municipio'],
            'total':  int(m.get('total', 0)),
            'm':      int(m.get('m', 0)),
            'h':      int(m.get('h', 0)),
        })

    # ── Apoyos (para tab Apoyos) — con árbol Apoyo > Inst > Prog ─────────────
    # Árbol desde desglose_municipal (mismo que reporte general sec. 6)
    apoyo_tree = {}
    for mun_k, entries in desglose_mun.items():
        for e in entries:
            ap_nom = e.get('apoyo', '')
            ins    = e.get('institucion', '')
            prog   = e.get('programa', '') or '(sin programa)'
            if not ap_nom or not ins:
                continue
            apoyo_tree.setdefault(ap_nom, {})
            apoyo_tree[ap_nom].setdefault(ins, {})
            apoyo_tree[ap_nom][ins].setdefault(prog, {'m': 0, 'h': 0, 'total': 0, 'muns': set()})
            apoyo_tree[ap_nom][ins][prog]['m']     += int(sf(e.get('m', 0)))
            apoyo_tree[ap_nom][ins][prog]['h']     += int(sf(e.get('h', 0)))
            apoyo_tree[ap_nom][ins][prog]['total'] += int(sf(e.get('total', 0)))
            apoyo_tree[ap_nom][ins][prog]['muns'].add(mun_k)

    # ── Rangos de edad por tipo de apoyo (suma desde desglose_municipal) ────────
    rangos_por_apoyo = {}
    _rk = ['0-5','6-11','12-17','18-29','30-49','50-64','65+','sin_datos']
    for mun_entries in desglose_mun.values():
        for entry in mun_entries:
            nombre_ap = entry.get('apoyo', '')
            if not nombre_ap:
                continue
            if nombre_ap not in rangos_por_apoyo:
                rangos_por_apoyo[nombre_ap] = {k: {'m':0,'h':0,'total':0} for k in _rk}
            for rk in _rk:
                rv = entry.get('rangos', {}).get(rk, {})
                rangos_por_apoyo[nombre_ap][rk]['m']     += sf(rv.get('m', 0))
                rangos_por_apoyo[nombre_ap][rk]['h']     += sf(rv.get('h', 0))
                rangos_por_apoyo[nombre_ap][rk]['total'] += sf(rv.get('total', 0))

    # ── Mapa de aliases: nombre en indicadores → nombre en desglose/hoja_s ─────
    IND_PROG_ALIAS = {
        'ATENCIÓN MÉDICA':    'MEDICHIHUAHUA',
        'ATENCION MEDICA':    'MEDICHIHUAHUA',
        'ESTANCIAS INFANTILES PARA EL DESARROLLO INTEGRAL DE LA NIÑEZ':
            'PROGAMA DE ESTANCIAS INFANTILES PARA EL DESARROLLO INTEGRAL DE LA NIÑEZ',
        'REHABILITACIÓN INTEGRAL Y APOYOS FUNCIONALES':
            'REHABILITACIÓN INTEGRAL FÍSICA Y APOYOS FUNCIONALES',
        'REHABILITACION INTEGRAL Y APOYOS FUNCIONALES':
            'REHABILITACIÓN INTEGRAL FÍSICA Y APOYOS FUNCIONALES',
        'ATENCIÓN A NIÑOS, NIÑAS, ADOLESCENTES Y JUVENTUDES':
            'ATENCIÓN A NIÑAS, NIÑOS, ADOLESCENTES Y JUVENTUDES',
        'ATENCION A NINOS, NINAS, ADOLESCENTES Y JUVENTUDES':
            'ATENCIÓN A NIÑAS, NIÑOS, ADOLESCENTES Y JUVENTUDES',
        # Incentivos nombre largo
        'INCENTIVOS ECONÓMICOS A ESTUDIANTES INDÍGENAS PARA SU PROFESIONALIZACIÓN':
            'Incentivos económicos a estudiantes indígenas para su profesionalización',
    }

    # ── Rangos ÚNICOS por programa — ya leídos en hoja_s_rangos arriba ───────
    if False: pass  # placeholder
    if _wb_s and 'Unicos y Rango de Edad' in _wb_s.sheetnames:
        ws_s  = _wb_s['Unicos y Rango de Edad']
        rows_s = list(ws_s.iter_rows(values_only=True))
        COL_N, COL_R0, COL_SIN, COL_TOT = 6, 7, 14, 15
        for row_s in rows_s:
            nombre_s = row_s[COL_N] if COL_N < len(row_s) else None
            if not isinstance(nombre_s, str) or not nombre_s.strip():
                continue
            nombre_s = nombre_s.strip()
            if nombre_s in ('M','H','Sin datos','BENEFICIARIOS UNICOS POR RANGO DE EDAD','BENEFICIARIOS ÚNICOS '):
                continue
            total_s = row_s[COL_TOT] if COL_TOT < len(row_s) else None
            if not isinstance(total_s, (int, float)):
                continue
            rangos_s = {}
            for i_r, k_r in enumerate(_rk2):
                v_r = row_s[COL_R0 + i_r] if (COL_R0 + i_r) < len(row_s) else None
                rangos_s[k_r] = int(v_r) if isinstance(v_r, (int, float)) and v_r else 0
            sin_d_s = row_s[COL_SIN] if COL_SIN < len(row_s) else None
            hoja_s_rangos[_nk(nombre_s)] = rangos_s
        hoja_s_sin[_nk(nombre_s)]    = int(sin_d_s) if isinstance(sin_d_s,(int,float)) and sin_d_s else 0

    # ── Municipios por programa (desde desglose) ──────────────────────────────
    ind_prog_muns = {}
    for mun_nombre, mun_entries in desglose_mun.items():
        for entry in mun_entries:
            prg = entry.get('programa', '').upper().strip()
            if not prg:
                continue
            if prg not in ind_prog_muns:
                ind_prog_muns[prg] = set()
            ind_prog_muns[prg].add(mun_nombre)

    # ── Municipios por apoyo (desde desglose_municipal) ────────────────────────
    municipios_por_apoyo = {}  # {apoyo: {mun: {m, h, total}}}
    for mun_nombre, mun_entries in desglose_mun.items():
        for entry in mun_entries:
            ap = entry.get('apoyo', '')
            if not ap:
                continue
            if ap not in municipios_por_apoyo:
                municipios_por_apoyo[ap] = {}
            if mun_nombre not in municipios_por_apoyo[ap]:
                municipios_por_apoyo[ap][mun_nombre] = {'m': 0, 'h': 0, 'total': 0}
            municipios_por_apoyo[ap][mun_nombre]['m']     += sf(entry.get('m', 0))
            municipios_por_apoyo[ap][mun_nombre]['h']     += sf(entry.get('h', 0))
            municipios_por_apoyo[ap][mun_nombre]['total'] += sf(entry.get('total', 0))

    # ── Rangos y municipios por programa (desde desglose_municipal) ─────────────
    _rk = ['0-5','6-11','12-17','18-29','30-49','50-64','65+','sin_datos']
    rangos_por_prog   = {}  # {(apoyo, programa): {rango: total}}
    municipios_por_prog = {}  # {(apoyo, programa): set(municipios)}

    for mun_nombre, mun_entries in desglose_mun.items():
        for entry in mun_entries:
            ap  = entry.get('apoyo', '')
            prg = entry.get('programa', '')
            if not ap or not prg:
                continue
            key = (ap, prg)
            if key not in rangos_por_prog:
                rangos_por_prog[key]    = {k: 0 for k in _rk}
                municipios_por_prog[key] = set()
            municipios_por_prog[key].add(mun_nombre)
            for rk in _rk:
                rv = entry.get('rangos', {}).get(rk, {})
                rangos_por_prog[key][rk] += sf(rv.get('total', 0) if isinstance(rv, dict) else rv)
    apoyos_data = []
    for a in ap_clean:
        nombre_apoyo = a.get('apoyo', '')
        inst_tree    = apoyo_tree.get(nombre_apoyo, {})
        insts = []
        for ins_k, prog_tree in sorted(inst_tree.items(),
                                       key=lambda x: -sum(v['total'] for v in x[1].values())):
            ins_total = sum(v['total'] for v in prog_tree.values())
            ins_m     = sum(v['m']     for v in prog_tree.values())
            ins_h     = sum(v['h']     for v in prog_tree.values())
            ins_muns  = set()
            for v in prog_tree.values():
                ins_muns |= v['muns']
            progs_list = []
            for prog_k, pv in sorted(prog_tree.items(), key=lambda x: -x[1]['total']):
                prog_key   = (nombre_apoyo, prog_k)
                rp_prog    = rangos_por_prog.get(prog_key, {})
                mp_prog    = municipios_por_prog.get(prog_key, set())
                _rkd       = ['0-5','6-11','12-17','18-29','30-49','50-64','65+']
                rangos_prog = {k: int(sf(rp_prog.get(k, 0))) for k in _rkd}
                muns_lista  = sorted(mp_prog)
                progs_list.append({
                    'nombre':     prog_k,
                    'total':      pv['total'],
                    'm':          pv['m'],
                    'h':          pv['h'],
                    'muns':       len(pv['muns']),
                    'muns_lista': muns_lista,
                    'rangos':     rangos_prog,
                })
            insts.append({
                'nombre': ins_k,
                'total':  ins_total,
                'm':      ins_m,
                'h':      ins_h,
                'muns':   len(ins_muns),
                'programas': progs_list,
            })
        # Rangos de edad para este apoyo
        _ra = rangos_por_apoyo.get(nombre_apoyo, {})
        _rkeys_data = ['0-5','6-11','12-17','18-29','30-49','50-64','65+']
        _rangos_apoyo = {k: int(sf(_ra.get(k, {}).get('total', 0))) for k in _rkeys_data}
        _keys_con_dato = [k for k in _rkeys_data if _rangos_apoyo[k] > 0]
        _rango_dom = max(_keys_con_dato, key=lambda k: _rangos_apoyo[k]) if _keys_con_dato else None
        _rango_min = min(_keys_con_dato, key=lambda k: _rangos_apoyo[k]) if len(_keys_con_dato) > 1 else None
        _rl = {'0-5':'0–5','6-11':'6–11','12-17':'12–17','18-29':'18–29','30-49':'30–49','50-64':'50–64','65+':'65+'}

        # Municipios con datos reales para este apoyo
        _muns_apoyo = municipios_por_apoyo.get(nombre_apoyo, {})
        _muns_sorted = sorted(_muns_apoyo.items(), key=lambda x: -x[1]['total'])
        _por_municipio = [
            {'nombre': mn, 'total': int(sv['total']), 'm': int(sv['m']), 'h': int(sv['h'])}
            for mn, sv in _muns_sorted
        ]

        apoyos_data.append({
            'nombre':        nombre_apoyo,
            'total':         int(sf(a.get('total', 0))),
            'm':             int(sf(a.get('m', 0))),
            'h':             int(sf(a.get('h', 0))),
            'n_muns':        int(sf(a.get('n_municipios', 0))),
            'pct':           pct(sf(a.get('total', 0)), total_apoyos),
            'instituciones': insts,
            'rangos':        _rangos_apoyo,
            'rango_dom':     _rl.get(_rango_dom, _rango_dom) if _rango_dom else None,
            'rango_min':     _rl.get(_rango_min, _rango_min) if _rango_min else None,
            'por_municipio': _por_municipio,
        })

    # ── Apoyos por institución (hoja 3) ──────────────────────────────────────
    apoyos_g3_summary = {}
    for inst_k, v in apoyos_g3.items():
        apoyos_g3_summary[inst_k] = {
            'total': int(sf(v.get('total', 0))),
            'm':     int(sf(v.get('m', 0))),
            'h':     int(sf(v.get('h', 0))),
        }

    # ── Indicadores y metas ───────────────────────────────────────────────────
    indicadores_data = []
    for ind in indicadores:
        def _int(v): return int(sf(v)) if v else None
        def _flt(v): return round(float(v), 2) if v else None
        nombre_prog_key   = ind.get('nombre', '').upper().strip()
        nombre_prog_alias = IND_PROG_ALIAS.get(nombre_prog_key,
                          IND_PROG_ALIAS.get(nombre_prog_key.title(), nombre_prog_key))
        _rk2 = ['0-5','6-11','12-17','18-29','30-49','50-64','65+']
        # Rangos desde hoja S (beneficiarios únicos, dato correcto)
        _rp  = hoja_s_rangos.get(_nk(nombre_prog_alias), hoja_s_rangos.get(_nk(nombre_prog_key), {}))
        _sin_datos_rango = hoja_s_sin.get(_nk(nombre_prog_alias), hoja_s_sin.get(_nk(nombre_prog_key), 0))
        # Municipios desde desglose
        _mp  = sorted(ind_prog_muns.get(nombre_prog_alias, ind_prog_muns.get(nombre_prog_key, set())))
        _clave = ind.get('clave', '') or ''
        # Limpiar clave inválida
        if not _clave or _clave in ('N/A','n/a','#DIV/0!','#N/A','') or _clave.startswith('#'):
            _clave = None
        indicadores_data.append({
            'inst':          ind.get('institucion', ''),
            'clave':         _clave,
            'nombre':        ind.get('nombre', ''),
            'pob_potencial': _int(ind.get('pob_potencial')),
            'pob_objetivo':  _int(ind.get('pob_objetivo')),
            'pob_alcanzada': _int(ind.get('pob_alcanzada')),
            'benef_unicos':  _int(ind.get('benef_unicos')),
            'benef_reales':  _int(ind.get('benef_reales')),
            'mujeres':       _int(ind.get('mujeres')),
            'hombres':       _int(ind.get('hombres')),
            'sin_id':        _int(ind.get('sin_id')),
            'presupuesto':   _flt(ind.get('presupuesto')),
            'gasto':         _flt(ind.get('gasto')),
            'ep':            _flt(ind.get('ep')),
            'metas_prog':    _flt(ind.get('metas_prog')),
            'avance_metas':  _flt(ind.get('avance_metas')),
            'eficacia':      _flt(ind.get('eficacia')),
            'eficiencia':    _flt(ind.get('eficiencia')),
            'desempeno':     _flt(ind.get('desempeño')),
            'rangos':        {k: int(_rp.get(k, 0)) for k in _rk2},
            'sin_datos_edad': _sin_datos_rango,
            'municipios':    _mp,
        })

    indicadores_data = []
    for ind in indicadores:
        def _int(v): return int(sf(v)) if v else None
        def _flt(v): return round(float(v), 2) if v else None
        nombre_prog_key = ind.get('nombre', '').upper().strip()
        nombre_prog_alias = IND_PROG_ALIAS.get(nombre_prog_key,
                          IND_PROG_ALIAS.get(nombre_prog_key.title(), nombre_prog_key))
        _rk2 = ['0-5','6-11','12-17','18-29','30-49','50-64','65+']
        # Rangos desde hoja S (beneficiarios únicos, dato correcto)
        _rp  = hoja_s_rangos.get(_nk(nombre_prog_alias), hoja_s_rangos.get(_nk(nombre_prog_key), {}))
        _sin_datos_rango = hoja_s_sin.get(_nk(nombre_prog_alias), hoja_s_sin.get(_nk(nombre_prog_key), 0))
        # Municipios desde desglose
        _mp  = sorted(ind_prog_muns.get(nombre_prog_alias, ind_prog_muns.get(nombre_prog_key, set())))
        _clave = ind.get('clave', '') or ''
        # Limpiar clave inválida
        if not _clave or _clave in ('N/A','n/a','#DIV/0!','#N/A','') or _clave.startswith('#'):
            _clave = None
        indicadores_data.append({
            'inst':          ind.get('institucion', ''),
            'clave':         _clave,
            'nombre':        ind.get('nombre', ''),
            'pob_potencial': _int(ind.get('pob_potencial')),
            'pob_objetivo':  _int(ind.get('pob_objetivo')),
            'pob_alcanzada': _int(ind.get('pob_alcanzada')),
            'benef_unicos':  _int(ind.get('benef_unicos')),
            'benef_reales':  _int(ind.get('benef_reales')),
            'mujeres':       _int(ind.get('mujeres')),
            'hombres':       _int(ind.get('hombres')),
            'sin_id':        _int(ind.get('sin_id')),
            'presupuesto':   _flt(ind.get('presupuesto')),
            'gasto':         _flt(ind.get('gasto')),
            'ep':            _flt(ind.get('ep')),
            'metas_prog':    _flt(ind.get('metas_prog')),
            'avance_metas':  _flt(ind.get('avance_metas')),
            'eficacia':      _flt(ind.get('eficacia')),
            'eficiencia':    _flt(ind.get('eficiencia')),
            'desempeno':     _flt(ind.get('desempeño')),
            'rangos':        {k: int(_rp.get(k, 0)) for k in _rk2},
            'sin_datos_edad': _sin_datos_rango,
            'municipios':    _mp,
        })

    # ── Presupuesto global (igual que reporte general) ────────────────────────
    pres_vals  = [float(p['presupuesto']) for p in indicadores if p.get('presupuesto') and float(p.get('presupuesto', 0)) > 0]
    gasto_vals = [float(p['gasto'])       for p in indicadores if p.get('gasto')       and float(p.get('gasto', 0)) > 0]
    pres_total  = sum(pres_vals)
    gasto_total = sum(gasto_vals)

    # ── Localizables rangos de edad ────────────────────────────────────────────
    loc_rangos_data = []
    for label, key in edad_labels:
        if key == 'sin_datos':
            continue
        t_e = int(sf(loc_rangos.get(key, 0)))
        loc_rangos_data.append({'label': label, 'key': key, 'total': t_e})

    # ── Grupos Vulnerables y NutriChihuahua ──────────────────────────────────
    grupos_vul   = leer_grupos_vulnerables(excel_path) if excel_path else {}
    nutrichi     = leer_nutrichihuahua(excel_path)     if excel_path else {}

    # Recalcular pob_vulnerable total desde Excel si hay datos reales
    gv_m   = grupos_vul.get('mujeres', {}).get('pob_vulnerable', 0) or 0
    gv_h   = grupos_vul.get('hombres', {}).get('pob_vulnerable', 0) or 0
    pob_vul_real = (gv_m + gv_h) if (gv_m + gv_h) > 0 else POB_VULNERABLE

    # ══ PAYLOAD FINAL ════════════════════════════════════════════════════════
    return {
        '_meta': {
            'pob_estatal':    POB_ESTATAL,
            'pob_vulnerable': pob_vul_real,
            'pob_vul_m':      gv_m,
            'pob_vul_h':      gv_h,
            'fuente':         'Padrón de Beneficiarios — SDHyBC Chihuahua',
        },
        # Reporte General
        'general': {
            'total_benef':   total_benef,
            'total_m':       total_m,
            'total_h':       total_h,
            'total_sn':      total_sn,
            'total_apoyos':  total_apoyos,
            'total_inst':    total_inst,
            'total_prog':    total_prog,
            'mun_activos':   67,
            'ninos':         ninos,
            'jovenes':       jovenes,
            'adultos':       adultos,
            'mayores':       mayores,
            'cob_estatal_pct':  pct(total_benef, POB_ESTATAL),
            'cob_vulnerable_pct': pct(total_benef, POB_VULNERABLE),
            'pres_total':    pres_total,
            'gasto_total':   gasto_total,
            'gasto_x_ben':   (gasto_total / total_benef) if total_benef and gasto_total else 0,
            'rangos_mh': {d['key']: {'m': d['m'], 'h': d['h'], 'total': d['total']}
                          for d in rangos_data if d['key'] != 'sin_datos'},
            'rangos':    {d['key']: d['total'] for d in rangos_data},
        },
        'rangos_edad': rangos_data,
        # Localizables
        'localizables': {
            'total': loc_total,
            'm':     loc_m,
            'h':     loc_h,
            'pct_del_padron': pct(loc_total, total_benef),
            'por_institucion': sorted(
                [{'nombre': x['nombre'], 'total': x['total'], 'm': x['m'], 'h': x['h']}
                 for x in loc_inst_raw],
                key=lambda x: -x['total']
            ),
            'por_municipio': loc_por_municipio,
            'rangos_edad':   loc_rangos_data,
        },
        # Instituciones
        'instituciones': instituciones_data,
        # Municipios
        'municipios': municipios_data,
        'municipios_especiales': mun_especiales,
        # Apoyos
        'apoyos': apoyos_data,
        'apoyos_g3': apoyos_g3_summary,
        # Indicadores
        'indicadores': indicadores_data,
        # Grupos Vulnerables (hoja nueva)
        'grupos_vulnerables': grupos_vul,
        # NutriChihuahua (hoja nueva)
        'nutrichihuahua': nutrichi,
    }


def main():
    if len(sys.argv) < 2:
        print('Uso: python3 generar_dashboard_data.py <excel_path> [--json]', file=sys.stderr)
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f'ERROR: no existe {excel_path}', file=sys.stderr)
        sys.exit(1)

    modo_json = '--json' in sys.argv

    print('Leyendo Excel...', file=sys.stderr)
    raw  = leer_excel(excel_path)
    print('Aplicando filtros...', file=sys.stderr)
    data = build_dashboard_data(raw, excel_path=excel_path)

    if modo_json:
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return

    # Generar data_dashboard.js junto al dashboard HTML
    out_js = SCRIPT_DIR / 'data_dashboard.js'
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    with open(out_js, 'w', encoding='utf-8') as f:
        f.write('// Generado automáticamente por generar_dashboard_data.py\n')
        f.write('// NO editar manualmente — se sobreescribe con cada actualización del Excel.\n')
        f.write(f'window.DASHBOARD_DATA = {payload};\n')

    kb = out_js.stat().st_size // 1024
    print(f'✓ data_dashboard.js generado ({kb} KB) → {out_js}', file=sys.stderr)
    print(f'  Beneficiarios únicos : {data["general"]["total_benef"]:,}', file=sys.stderr)
    print(f'  Apoyos otorgados     : {data["general"]["total_apoyos"]:,}', file=sys.stderr)
    print(f'  Municipios activos   : {data["general"]["mun_activos"]}', file=sys.stderr)
    print(f'  Instituciones activas: {data["general"]["total_inst"]}', file=sys.stderr)
    print(f'  Localizables         : {data["localizables"]["total"]:,}', file=sys.stderr)
    if data.get('grupos_vulnerables'):
        gv = data['grupos_vulnerables']
        print(f'  Pob. Vul. Mujeres    : {gv.get("mujeres",{}).get("pob_vulnerable",0):,}', file=sys.stderr)
        print(f'  Pob. Vul. Hombres    : {gv.get("hombres",{}).get("pob_vulnerable",0):,}', file=sys.stderr)
    if data.get('nutrichihuahua') and data['nutrichihuahua']:
        print(f'  NutriChihuahua       : {len(data["nutrichihuahua"])} registros', file=sys.stderr)


if __name__ == '__main__':
    main()
