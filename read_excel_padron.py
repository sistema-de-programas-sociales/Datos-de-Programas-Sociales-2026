#!/usr/bin/env python3
"""
read_excel_padron.py — Lector completo del padrón de beneficiarios Chihuahua

Jerarquía sheet 3 "Apoyos Otorgados":
  APOYO (row donde siguiente es INST)
    INST  (nombre de institución conocida)
      PROG  (primer no-age post-INST, antes de que llegue un MUN)
        MUN   (municipio: no es AGE, no es INST, no es PROG — viene después del PROG)
          AGE   (rangos de edad: 0-5, 6-11, ... → M, H, Total)
"""
import openpyxl, json, sys
from pathlib import Path

EXCEL = sys.argv[1]
wb    = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)

# ── CACHÉ DE HOJAS — se cargan una sola vez en memoria ───────────────────────
# Elimina 80+ re-lecturas del archivo Excel durante calcular_filtro
_WB_CACHE = {sh: list(wb[sh].iter_rows(values_only=True)) for sh in wb.sheetnames}

# ── Utilidades ────────────────────────────────────────────────────────────────
def sf(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def clean(v):
    if v is None: return None
    s = str(v).strip(); return s if s else None

# ── Correcciones ortográficas de datos del Excel ─────────────────────────────
# Corrige tildes y errores tipográficos en apoyos, programas y municipios
_CORRECCIONES = {
    # Apoyos
    "ACOMPAÑAMIENTO EN SITUACION EMERGENTE":   "ACOMPAÑAMIENTO EN SITUACIÓN EMERGENTE",
    "APOYO ECONOMICO":                          "APOYO ECONÓMICO",
    "APOYOS MEDICOS":                           "APOYOS MÉDICOS",
    "APOYOS MEDICOS - CUENTA HOSPITALARIA":     "APOYOS MÉDICOS - CUENTA HOSPITALARIA",
    "APOYOS MEDICOS - DIALISIS":                "APOYOS MÉDICOS - DIÁLISIS",
    "APOYOS MEDICOS - ESTUDIO DE LABORATORIO":  "APOYOS MÉDICOS - ESTUDIO DE LABORATORIO",
    "APOYOS MEDICOS - IMAGENOLOGIA":            "APOYOS MÉDICOS - IMAGENOLOGÍA",
    "APOYOS MEDICOS - INSUMOS MEDICOS":         "APOYOS MÉDICOS - INSUMOS MÉDICOS",
    "APOYOS MEDICOS - MATERIAL DE CURACION":    "APOYOS MÉDICOS - MATERIAL DE CURACIÓN",
    "APOYOS MEDICOS - MEDICAMENTO":             "APOYOS MÉDICOS - MEDICAMENTO",
    "APOYOS MEDICOS - OTROS TIPOS":             "APOYOS MÉDICOS - OTROS TIPOS",
    "APOYOS MEDICOS - PAÑALES":                 "APOYOS MÉDICOS - PAÑALES",
    "APOYOS MEDICOS- COLOSTOMIA":               "APOYOS MÉDICOS - COLOSTOMÍA",
    "CAPACITACION EN EMBARAZO Y LACTANCIA":     "CAPACITACIÓN EN EMBARAZO Y LACTANCIA",
    "CAPACITACION EN ORIENTACION ALIMENTARIA":  "CAPACITACIÓN EN ORIENTACIÓN ALIMENTARIA",
    "CONDONACION O ESTIMULO FISCAL":            "CONDONACIÓN O ESTÍMULO FISCAL",
    "CONSULTA MEDICA ESPECIALIADA EN REHABILITACION FISICA":
                                                "CONSULTA MÉDICA ESPECIALIZADA EN REHABILITACIÓN FÍSICA",
    "CONSULTA MEDICA LOCAL":                    "CONSULTA MÉDICA LOCAL",
    "CONSULTA MEDICA PARA REHAB. CARDIOVASCULAR Y RESPIRATORIO":
                                                "CONSULTA MÉDICA PARA REHAB. CARDIOVASCULAR Y RESPIRATORIO",
    "CULTURA Y RECREACION.":                    "CULTURA Y RECREACIÓN",
    "CURSO DE CAPACITACION":                    "CURSO DE CAPACITACIÓN",
    "EDUCACION Y CAPACITACION TECNICA":         "EDUCACIÓN Y CAPACITACIÓN TÉCNICA",
    "EVENTOS LUDICOS":                          "EVENTOS LÚDICOS",
    "HOSPEDAJE Y ALIMENTACION ESTANCIA TEMPORAL":
                                                "HOSPEDAJE Y ALIMENTACIÓN - ESTANCIA TEMPORAL",
    "PRECONSULTA MEDICA":                       "PRECONSULTA MÉDICA",
    "PREVENCION DE LA VIOLENCIA SEXUAL EN NNA": "PREVENCIÓN DE LA VIOLENCIA SEXUAL EN NNA",
    "PRUEBAS DE DIAGNOSTICO -TRASTORNO DE ESPECTRO AUSTISTA (TEA)":
                                                "PRUEBAS DE DIAGNÓSTICO - TRASTORNO DEL ESPECTRO AUTISTA (TEA)",
    "TERAPIA NEUROSENSORIAL Y ESTIMULACION TEMPRANA":
                                                "TERAPIA NEUROSENSORIAL Y ESTIMULACIÓN TEMPRANA",
    "TRAMITE FUNERARIO APOYADO":                "TRÁMITE FUNERARIO APOYADO",
    "TRAMITE FUNERARIO GESTIONADO":             "TRÁMITE FUNERARIO GESTIONADO",
    "VALORACION NUTRICIONAL A MUJER EMBARAZDA": "VALORACIÓN NUTRICIONAL A MUJER EMBARAZADA",
    "VALORACION NUTRICIONAL A NIÑOS Y NIÑAS":   "VALORACIÓN NUTRICIONAL A NIÑOS Y NIÑAS",
    # Programas
    "ASISTENCIA SOCIAL PARA LA POBLACION INDIGENA":
                                                "ASISTENCIA SOCIAL PARA LA POBLACIÓN INDÍGENA",
    "ATENCION A LA JUVENTUD":                   "ATENCIÓN A LA JUVENTUD",
    "ATENCION A NIÑAS, NIÑOS, ADOLESCENTES Y JUVENTUDES":
                                                "ATENCIÓN A NIÑAS, NIÑOS, ADOLESCENTES Y JUVENTUDES",
    "FORTALECIMIENTO COMUNITARIO Y PARTICIPACION CIUDADANA":
                                                "FORTALECIMIENTO COMUNITARIO Y PARTICIPACIÓN CIUDADANA",
    "GESTION SOCIAL Y ATENCION A LA CIUDADANIA":
                                                "GESTIÓN SOCIAL Y ATENCIÓN A LA CIUDADANÍA",
    "APOYOS A PERSONAS A DISCAPACIDAD":         "APOYOS A PERSONAS CON DISCAPACIDAD",
    # Programas sin tilde
    "ALIMENTACION Y DESARROLLO AUTOSUSTENTABLE DE LAS FAMILIAS":
                                                "ALIMENTACIÓN Y DESARROLLO AUTOSUSTENTABLE DE LAS FAMILIAS",
    "JUAREZ CUENTA CONMIGO":                    "JUÁREZ CUENTA CONMIGO",
    "PROYECTOS PRODUCTIVOS Y ECONOMIA SOLIDARIA":
                                                "PROYECTOS PRODUCTIVOS Y ECONOMÍA SOLIDARIA",
    "REHABILITACION INTEGRAL FISICA Y APOYOS FUNCIONALES":
                                                "REHABILITACIÓN INTEGRAL FÍSICA Y APOYOS FUNCIONALES",
    # Apoyos sin tilde
    "APOYOS A PERSONAS CON DISCAPACIDAD - ANDADOR":
                                                "APOYOS A PERSONAS CON DISCAPACIDAD - ANDADOR",
    "APOYOS DE DEFUNCION - SERVICIO FUNERARIO": "APOYOS DE DEFUNCIÓN - SERVICIO FUNERARIO",
    "APOYOS ESCOLARES - UTILES ESCOLARES":      "APOYOS ESCOLARES - ÚTILES ESCOLARES",
    "APOYOS PARA PERSONAS CON DISCAPACIDAD - APARATO AUDITIVO":
                                                "APOYOS PARA PERSONAS CON DISCAPACIDAD - APARATO AUDITIVO",
    "ASISTENCIA ALIMENTARIA EN ESPACIO COMUN":  "ASISTENCIA ALIMENTARIA EN ESPACIO COMÚN",
    "CURSO DE NATACION":                        "CURSO DE NATACIÓN",
    "DOCUMENTOS DE GESTION INSTITUCIONAL":      "DOCUMENTOS DE GESTIÓN INSTITUCIONAL",
    "GESTION DE APOYO":                         "GESTIÓN DE APOYO",
    "MODIFICACION CONDUCTUAL":                  "MODIFICACIÓN CONDUCTUAL",
    "PROTESIS":                                 "PRÓTESIS",
    "REPARACION Y AJUSTE DE ORTESIS":           "REPARACIÓN Y AJUSTE DE ÓRTESIS",
    "REPARACION Y AJUSTE DE PROTESIS":          "REPARACIÓN Y AJUSTE DE PRÓTESIS",
    "TALLER DE CAPACITACION":                   "TALLER DE CAPACITACIÓN",
    "TERAPIA FISICA":                           "TERAPIA FÍSICA",
    "TERAPIA PSICOLOGICA":                      "TERAPIA PSICOLÓGICA",
    # Municipios
    "AQUILES SERDAN":                   "AQUILES SERDÁN",
    "ASCENSION":                        "ASCENSIÓN",
    "BATOPILAS DE MANUEL GOMEZ MORIN":  "BATOPILAS DE MANUEL GÓMEZ MORÍN",
    "CUAUHTEMOC":                       "CUAUHTÉMOC",
    "FORANEO":                          "FORÁNEO",
    "GOMEZ FARIAS":                     "GÓMEZ FARÍAS",
    "HUEJOTITAN":                       "HUEJOTITÁN",
    "JIMENEZ":                          "JIMÉNEZ",
    "JUAREZ":                           "JUÁREZ",
    "LOPEZ":                            "LÓPEZ",
    "PRAXEDIS G. GUERRERO":             "PRAXÉDIS G. GUERRERO",
    "TEMOSACHIC":                       "TEMÓSACHIC",
}

def corr(s):
    """Aplica correcciones ortográficas si el texto está en el logo."""
    if not s: return s
    return _CORRECCIONES.get(s.strip(), s)

def _norm(s):
    import unicodedata
    return unicodedata.normalize('NFD', str(s).upper()).encode('ascii', 'ignore').decode()


INST_SET = {'CECYTECH','COESPO','COESVI','DIF','ICHD','ICHDII','ICHIJUV','ICHIMUJ',
            'RURAL','SALUD','SDBYBC','SDHyBC','SDHYBC','SEECH','SEYD','SEyD',
            'SPyCI','SPYCI','TRABAJO','TURISMO','CULTURA'}

AGE_KEYS_RAW = {'0-5','6-11','12-17','18-29','30-49','50-64','65+',
                'SIN DATOS','SIN DATO','SNDATOS','SINDATOS','SIN DATOS ',
                'Sin datos','M','H'}

SKIP_ROWS = {'RECUENTO DE # DE APOYOS',
             'NUMERO DE BENEFICIARIOS UNICOS POR MUNICIPIO',
             'BENEFICIARIOS UNICOS POR SEXO',
             'BENEFICIARIOS UNICOS POR RANGO DE EDAD',
             'BENEFICIARIOS ÚNICOS','BENEFICIARIOS UNICOS','D',''}

EXCLUIR_MUN = {'SIN DATOS','TOTAL','D',''}
ESPECIALES  = {'NO IDENTIFICADO','FORANEO','FORÁNEO'}

AGE_NORM = {
    '0-5':'0-5','6-11':'6-11','12-17':'12-17','18-29':'18-29',
    '30-49':'30-49','50-64':'50-64','65+':'65+',
    'SIN DATOS':'sin_datos','SIN DATO':'sin_datos','SNDATOS':'sin_datos',
    'SINDATOS':'sin_datos','SIN DATOS ':'sin_datos','SIN DATO ':'sin_datos',
    'Sin datos':'sin_datos',
}


POB_MUNICIPAL = {
    'AHUMADA':16198,'ALDAMA':27591,'ALLENDE':8403,
    'AQUILES SERDAN':33187,'AQUILES SERDÁN':33187,
    'ASCENSION':27978,'ASCENSIÓN':27978,
    'BACHINIVA':5850,'BALLEZA':16406,
    'BATOPILAS DE MANUEL GOMEZ MORIN':11069,'BATOPILAS':11069,
    'BOCOYNA':23060,'BUENAVENTURA':27426,
    'CAMARGO':48426,'CARICHI':7969,'CARICHÍ':7969,
    'CASAS GRANDES':12513,'CHIHUAHUA':1028306,
    'CHINIPAS':5960,'CHÍNIPAS':5960,
    'CORONADO':2060,'COYAME DEL SOTOL':1218,'COYAME':1218,
    'CUAUHTEMOC':196633,'CUAUHTÉMOC':196633,
    'CUSIHUIRIACHI':5826,'DELICIAS':156678,
    'DR. BELISARIO DOMINGUEZ':2475,'DR BELISARIO DOMINGUEZ':2475,
    'EL TULE':1369,'GALEANA':7291,
    'GOMEZ FARIAS':6778,'GÓMEZ FARÍAS':6778,
    'GRAN MORELOS':2484,'GUACHOCHI':56871,
    'GUADALUPE':3708,'GUADALUPE Y CALVO':50243,
    'GUAZAPARES':9305,'GUERRERO':34977,
    'HIDALGO DEL PARRAL':127636,
    'HUEJOTITAN':787,'HUEJOTITÁN':787,
    'IGNACIO ZARAGOZA':5040,'JANOS':11321,
    'JIMENEZ':39683,'JIMÉNEZ':39683,
    'JUAREZ':1661295,'JUÁREZ':1661295,
    'JULIMES':5734,'LA CRUZ':3686,
    'LOPEZ':4291,'LÓPEZ':4291,
    'MADERA':24000,'MAGUARICHI':1277,
    'MANUEL BENAVIDES':1103,
    'MATACHI':2700,'MATACHÍ':2700,
    'MATAMOROS':4351,'MEOQUI':46611,
    'MORELOS':7331,'MORIS':4478,
    'NAMIQUIPA':22649,'NONOAVA':3036,
    'NUEVO CASAS GRANDES':68506,'OCAMPO':8965,
    'OJINAGA':24243,
    'PRAXEDIS G. GUERRERO':4842,'PRAXEDIS GUERRERO':4842,
    'RIVA PALACIO':7722,'ROSALES':17031,
    'ROSARIO':2196,
    'SAN FRANCISCO DE BORJA':2315,
    'SAN FRANCISCO DE CONCHOS':3030,
    'SAN FRANCISCO DEL ORO':5027,
    'SANTA BARBARA':12579,'SANTA BÁRBARA':12579,
    'SANTA ISABEL':3814,
    'SATEVO':3793,'SATEVÓ':3793,
    'SAUCILLO':29693,
    'TEMOSACHIC':5241,'TEMÓSACHIC':5241,
    'URIQUE':16988,'URUACHI':7151,
    'VALLE DE ZARAGOZA':4727,
}
POB_ESTATAL = 4043130  # Suma real de los 67 municipios

def is_inst(n): return bool(n and n.strip().upper() in {i.upper() for i in INST_SET})
def is_age(n):  return bool(n and n.strip() in AGE_NORM or (n and n.strip().upper() in {a.upper() for a in AGE_KEYS_RAW}))
# Municipios válidos: índice normalizado (sin tildes) de los 67 municipios + especiales
_POB_NORM = {_norm(k): v for k, v in POB_MUNICIPAL.items()}  # lookup sin tildes
_MUN_VALIDOS_NORM = set(_POB_NORM.keys()) | {_norm(e) for e in {'NO IDENTIFICADO','FORANEO','FORÁNEO'}}
def is_municipio(n): return bool(n and _norm(n.strip()) in _MUN_VALIDOS_NORM)
def is_skip(n): return bool(n and n.strip().upper() in {s.upper() for s in SKIP_ROWS})

# Programa que bypasea el filtro de threshold (tiene beneficiarios reales aunque total < 10)
_EXCEPCION_PROG = 'ATENCIÓN A NIÑAS, NIÑOS, ADOLESCENTES Y JUVENTUDES'  # forma corregida

def es_placeholder(sn, m, h, total, nombre=''):
    """
    Excluye programas/instituciones con menos de 10 beneficiarios donde
    todos son sin dato de sexo (m=0, h=0). Esto cubre valores 0-9.
    Excepción: el programa de Atención a Niñas, Niños, Adolescentes y Juventudes nunca se excluye.
    """
    if nombre and _norm(nombre.strip()) == _norm(_EXCEPCION_PROG):
        return False
    return sf(total) < 10 and sf(m) == 0 and sf(h) == 0

def norm_age(n):
    if n is None: return None
    s = n.strip()
    if s in AGE_NORM: return AGE_NORM[s]
    su = s.upper()
    for k,v in AGE_NORM.items():
        if k.upper() == su: return v
    return None

def empty_rmh():
    return {k: {'m':0.0,'h':0.0,'total':0.0}
            for k in ['0-5','6-11','12-17','18-29','30-49','50-64','65+','sin_datos']}

def add_rmh(dst, age_key, m, h, t):
    dst[age_key]['m']     += m
    dst[age_key]['h']     += h
    dst[age_key]['total'] += t

def rmh_to_simple(rmh):
    return {k: v['total'] for k, v in rmh.items()}

# ─── SHEET 1: Únicos y Rango de Edad ─────────────────────────────────────────
def parse_sheet1():
    rows = [r for r in _WB_CACHE.get('Unicos y Rango de Edad', []) if any(c is not None for c in r)]
    instituciones = {}; rangos_global = {}; gran_total = {}; inst_act = None
    # Estado para lectura de filas M/H en Tabla B (col G)
    _tab_b_inst = None; _tab_b_prog = None

    for r in rows:
        na = clean(r[0])
        if na and not is_skip(na) and not is_age(na):
            if na.upper() == 'TOTAL':
                gran_total = {'sn':sf(r[1]),'m':sf(r[2]),'h':sf(r[3]),'total':sf(r[4])}
            elif is_inst(na):
                inst_act = na
                if na not in instituciones:
                    instituciones[na] = {'sn':sf(r[1]),'m':sf(r[2]),'h':sf(r[3]),'total':sf(r[4]),'programas':[],'rangos':{}}
                else:
                    instituciones[na]['sn']    += sf(r[1])
                    instituciones[na]['m']     += sf(r[2])
                    instituciones[na]['h']     += sf(r[3])
                    instituciones[na]['total'] += sf(r[4])
            elif inst_act:
                na_corr = corr(na)
                t_prog = sf(r[4])
                if es_placeholder(r[1], r[2], r[3], r[4], na_corr): continue
                if t_prog >= 10 or _norm(na_corr) == _norm(_EXCEPCION_PROG):
                    instituciones[inst_act]['programas'].append(
                        {'nombre':na_corr,'sn':sf(r[1]),'m':sf(r[2]),'h':sf(r[3]),'total':t_prog})

        ng = clean(r[6])
        if ng and not is_skip(ng):
            # ── Detectar filas M/H de la Tabla B ─────────────────────────────
            # La tabla dinámica tiene: INST → PROG → M → H (filas de sexo)
            # Cuando ng == 'M' o 'H', pertenece al último programa visto en col G
            _ng_u = ng.strip().upper()
            if _ng_u in ('M', 'H'):
                # Acumular en el último programa de la institución activa en TabB
                if _tab_b_inst and _tab_b_prog is not None:
                    _rng_fila = {'0-5':sf(r[7]),'6-11':sf(r[8]),'12-17':sf(r[9]),
                                 '18-29':sf(r[10]),'30-49':sf(r[11]),'50-64':sf(r[12]),
                                 '65+':sf(r[13]),'sin_datos':sf(r[14])}
                    sexo = _ng_u.lower()  # 'm' o 'h'
                    if _tab_b_inst in instituciones:
                        progs = instituciones[_tab_b_inst]['programas']
                        # Buscar el programa por nombre en la lista
                        for p in progs:
                            if _norm(p['nombre']) == _norm(_tab_b_prog):
                                if 'rangos_mh' not in p:
                                    p['rangos_mh'] = {}
                                p['rangos_mh'][sexo] = _rng_fila
                                break
                continue

            if is_age(ng):
                continue

            rng = {'0-5':sf(r[7]),'6-11':sf(r[8]),'12-17':sf(r[9]),
                   '18-29':sf(r[10]),'30-49':sf(r[11]),'50-64':sf(r[12]),
                   '65+':sf(r[13]),'sin_datos':sf(r[14])}
            if ng.upper() == 'TOTAL':
                rangos_global = rng
                _tab_b_inst = None; _tab_b_prog = None
            elif is_inst(ng):
                _tab_b_inst = ng; _tab_b_prog = None
                if ng in instituciones:
                    instituciones[ng]['rangos'] = rng
            elif _tab_b_inst:
                # Fila de programa en Tabla B — guardar nombre para asociar M/H siguientes
                _tab_b_prog = corr(ng)

    # ── Tabla S: desglose M/H por rango de edad ──────────────────────────────
    # La tabla real está en col S (idx 18) de la hoja "Unicos y Rango de Edad".
    # Header en la fila que contiene 'BENEFICIARIOS UNICOS POR RANGO DE EDAD' en col 18.
    # Subcolumnas: col 19 = Sin datos, col 20 = M, col 21 = H, col 22 = TOTAL.
    AGE_KEY_MAP = {'0-5':'0-5','6-11':'6-11','12-17':'12-17','18-29':'18-29',
                   '30-49':'30-49','50-64':'50-64','65+':'65+','sin datos':'sin_datos'}
    rangos_mh_global = {}
    all_rows_s = list(_WB_CACHE.get('Unicos y Rango de Edad', []))

    # 1. Localizar la fila de encabezado buscando en col 18 (col S)
    hdr_idx_s = None
    for ri, row in enumerate(all_rows_s):
        v = row[18] if len(row) > 18 else None
        if v and 'BENEFICIARIOS UNICOS POR RANGO DE EDAD' in str(v).upper():
            hdr_idx_s = ri
            break

    if hdr_idx_s is not None:
        # Col fija: 18=label, 19=Sin datos, 20=M, 21=H, 22=TOTAL
        # (autodescubrimiento defensivo por si el layout cambia)
        hdr_row_s = all_rows_s[hdr_idx_s]
        col_label_s = 18
        col_sn_s = col_m_s = col_h_s = col_t_s = None
        for ci in range(col_label_s + 1, min(col_label_s + 8, len(hdr_row_s))):
            h = str(hdr_row_s[ci] or '').strip().upper()
            if h in ('SIN DATOS', 'SIN DATO') and col_sn_s is None:
                col_sn_s = ci
            elif h == 'M' and col_m_s is None:
                col_m_s = ci
            elif h == 'H' and col_h_s is None:
                col_h_s = ci
            elif h == 'TOTAL' and col_t_s is None:
                col_t_s = ci
        # fallback a posiciones fijas si el header no fue detectado
        if col_sn_s is None: col_sn_s = 19
        if col_m_s  is None: col_m_s  = 20
        if col_h_s  is None: col_h_s  = 21
        if col_t_s  is None: col_t_s  = 22

        # 2. Leer filas de datos hasta encontrar TOTAL
        for row_s in all_rows_s[hdr_idx_s + 1:]:
            label = str(row_s[col_label_s] if len(row_s) > col_label_s else '').strip().lower()
            if label == 'total':
                break
            key = AGE_KEY_MAP.get(label)
            if key:
                rangos_mh_global[key] = {
                    'sn':    int(sf(row_s[col_sn_s] if len(row_s) > col_sn_s else None)),
                    'm':     int(sf(row_s[col_m_s]  if len(row_s) > col_m_s  else None)),
                    'h':     int(sf(row_s[col_h_s]  if len(row_s) > col_h_s  else None)),
                    'total': int(sf(row_s[col_t_s]  if len(row_s) > col_t_s  else None)),
                }

    inst_filtradas = {k: v for k, v in instituciones.items()
                      if not es_placeholder(v['sn'], v['m'], v['h'], v['total'], k)}
    return inst_filtradas, rangos_global, gran_total, rangos_mh_global

# ─── SHEET 2: Beneficiarios por Municipio ────────────────────────────────────
def find_block_positions(non_empty):
    """Filas donde la SIGUIENTE fila no-skip es una INSTITUCIÓN → esa fila es el encabezado del bloque."""
    pos = []
    for i, r in enumerate(non_empty[:-1]):
        n = clean(r[0])
        if not n or is_skip(n) or is_age(n) or is_inst(n): continue
        if n.upper() in ('TOTAL',): continue
        for j in range(i+1, min(i+3, len(non_empty))):
            nn = clean(non_empty[j][0])
            if not nn or is_skip(nn): continue
            if is_inst(nn): pos.append(i)
            break
    return pos

def parse_sheet2():
    rows = [r for r in _WB_CACHE.get('Beneficiarios por Municipio', []) if any(c is not None for c in r)]
    mun_positions = find_block_positions(rows)
    municipios = []

    for idx, pos in enumerate(mun_positions):
        r_mun  = rows[pos]
        nombre = corr(clean(r_mun[0]))
        if nombre.upper() in EXCLUIR_MUN: continue
        es_especial = _norm(nombre) in {_norm(e) for e in ESPECIALES}
        m = sf(r_mun[2]); h = sf(r_mun[3]); total = sf(r_mun[4])
        sn = sf(r_mun[1])
        end = mun_positions[idx+1] if idx+1 < len(mun_positions) else len(rows)

        rangos_mh = empty_rmh()
        progs_detail = []
        insts = set()
        inst_act = None; prog_act = None
        sf_psn = sf_pm = sf_ph = sf_pt = 0
        prog_rmh = None
        # Subtotales por institución en este municipio (fila de institución en el Excel)
        inst_subtotales = {}  # {inst_key: {m, h, total}}

        def flush():
            if prog_act is not None:
                if es_placeholder(sf_psn, sf_pm, sf_ph, sf_pt, prog_act): return
                progs_detail.append({
                    'nombre':prog_act,'institucion':inst_act,
                    'sn':sf_psn,'m':sf_pm,'h':sf_ph,'total':sf_pt,
                    'rangos':dict(prog_rmh),
                })
                insts.add(inst_act)

        for r in rows[pos+1:end]:
            n  = clean(r[0])
            if n is None: continue
            ak = norm_age(n)
            if ak:
                am=sf(r[2]); ah=sf(r[3]); at=sf(r[4])
                add_rmh(rangos_mh, ak, am, ah, at)
                if prog_rmh is not None: add_rmh(prog_rmh, ak, am, ah, at)
                continue
            if is_inst(n):
                flush(); inst_act=n; insts.add(n); prog_act=None; prog_rmh=None
                # Guardar subtotal de la fila de institución (ya deduplica el Excel)
                inst_subtotales[n] = {'m': sf(r[2]), 'h': sf(r[3]), 'total': sf(r[4])}
                sf_pm=sf_ph=sf_pt=0; continue
            if inst_act:
                flush()
                prog_act=corr(n); sf_psn=sf(r[1]); sf_pm=sf(r[2]); sf_ph=sf(r[3]); sf_pt=sf(r[4])
                prog_rmh=empty_rmh(); continue
        flush()

        if total > 0:
            municipios.append({
                'municipio':       nombre,
                'sn':sn,'m':m,'h':h,'total':total,
                'n_programas':     len(progs_detail),
                'n_instituciones': len(insts),
                'programas':       sorted({p['nombre'] for p in progs_detail}),
                'programas_detail':progs_detail,
                'instituciones':   sorted(insts),
                'rangos':          rmh_to_simple(rangos_mh),
                'rangos_mh':       rangos_mh,
                'inst_subtotales':  inst_subtotales,
                'poblacion':       _POB_NORM.get(_norm(nombre), 0),
                'especial':        es_especial,
            })
    municipios.sort(key=lambda x: -x['total'])
    return municipios

# ─── SHEET 3: Apoyos Otorgados ────────────────────────────────────────────────
def parse_sheet3_full():
    """
    Parser CORRECTO usando la jerarquía real:
      APOYO → INST → PROG → MUN(m,h,total) → AGE(m,h,total)

    Detección:
    - APOYO: fila no-inst, no-age, no-skip donde el siguiente no-skip es una INST
    - INST:  nombre en INST_SET
    - PROG:  primera fila no-inst, no-age después de INST
    - MUN:   filas no-inst, no-age DESPUÉS del PROG (puede haber varios MUN)
    - AGE:   filas con nombre en AGE_KEYS, siempre pertenecen al MUN anterior
    """
    rows = [r for r in _WB_CACHE.get('Apoyos Otorgados', []) if any(c is not None for c in r)]

    # ── pre-pasar: etiquetar cada fila ──────────────────────────────────────
    # Necesitamos saber para cada fila: ¿es APOYO?
    # Una fila es APOYO si su siguiente fila no-skip-no-age es una INST.
    tagged = []
    skip_set = {s.upper() for s in SKIP_ROWS}

    def next_real(i):
        for j in range(i+1, len(rows)):
            n = clean(rows[j][0])
            if n and n.upper() not in skip_set and not is_age(n):
                return n
        return None

    for i, r in enumerate(rows):
        n = clean(r[0])
        if not n or n.upper() in skip_set:
            tagged.append(('SKIP', n, r)); continue
        if is_age(n):
            tagged.append(('AGE', n, r)); continue
        if is_inst(n):
            tagged.append(('INST', n, r)); continue
        nr = next_real(i)
        if nr and is_inst(nr):
            n_u = n.strip().upper()
            ESPECIALES_APOYO = {'NO IDENTIFICADO', 'FORANEO', 'FORÁNEO'}
            # Municipios reales (67) → nunca son APOYOs
            # Especiales (NO IDENTIFICADO, FORÁNEO) → APOYO cuando next es INST
            # Cualquier otra cosa no-municipio → APOYO cuando next es INST
            if n_u in ESPECIALES_APOYO or _norm(n) not in _MUN_VALIDOS_NORM:
                tagged.append(('APOYO', n, r)); continue
        tagged.append(('OTHER', n, r))

    # ── construir desglose ──────────────────────────────────────────────────
    # Resultado global para apoyos list (sección 5 del reporte)
    apoyos_global = {}   # {apoyo_nombre: {m,h,total,n_municipios:set}}

    # Desglose por municipio: {mun: {apoyo: {m,h,total,rangos_mh,inst,prog}}}
    desglose = {}

    apoyo_act = None; inst_act = None; prog_act = None; mun_act = None

    for tag, name, r in tagged:
        if tag == 'SKIP': continue

        if tag == 'APOYO':
            apoyo_act = corr(name); inst_act = None; prog_act = None; mun_act = None
            # Inicializar entrada global
            if apoyo_act not in apoyos_global:
                apoyos_global[apoyo_act] = {'m':0,'h':0,'total':0,'municipios':set(),'programas':set()}
            continue

        if tag == 'INST':
            inst_act = name; prog_act = None; mun_act = None; continue

        if tag == 'AGE':
            if mun_act and apoyo_act and _norm(mun_act) in desglose:
                ak = norm_age(name)
                if ak:
                    desg_key = f"{apoyo_act}|||{inst_act}|||{prog_act}"
                    if desg_key in desglose[_norm(mun_act)]:
                        add_rmh(desglose[_norm(mun_act)][desg_key]['rangos'], ak, sf(r[2]), sf(r[3]), sf(r[4]))
            continue

        # tag == 'OTHER'
        if apoyo_act is None: continue

        if inst_act and not is_municipio(name):
            # No es municipio → es un PROGRAMA (nuevo o primero bajo esta inst)
            prog_act = corr(name); mun_act = None
            if apoyo_act and apoyo_act in apoyos_global:
                apoyos_global[apoyo_act]['programas'].add(prog_act)
            continue

        if inst_act and prog_act is not None and is_municipio(name):
            # OTHER después de PROG → es un MUNICIPIO
            mun = corr(name)
            mv = sf(r[2]); hv = sf(r[3]); tv = sf(r[4])
            if mun.upper() in EXCLUIR_MUN: continue

            mun_act = mun

            # Acumular en global — excluir especiales del conteo de municipios reales
            apoyos_global[apoyo_act]['m']     += mv
            apoyos_global[apoyo_act]['h']     += hv
            apoyos_global[apoyo_act]['total'] += tv
            if _norm(mun) not in {_norm(e) for e in ESPECIALES}:
                apoyos_global[apoyo_act]['municipios'].add(mun)

            # Acumular en desglose municipal (una fila por apoyo+inst+programa)
            mun_dk = _norm(mun)
            if mun_dk not in desglose: desglose[mun_dk] = {}
            desg_key = f"{apoyo_act}|||{inst_act}|||{prog_act}"
            if desg_key not in desglose[mun_dk]:
                desglose[mun_dk][desg_key] = {
                    'apoyo':apoyo_act, 'institucion':inst_act, 'programa':prog_act,
                    'm':0,'h':0,'total':0, 'rangos':empty_rmh(),
                }
            e = desglose[mun_dk][desg_key]
            e['m']     += mv
            e['h']     += hv
            e['total'] += tv
            continue

    # ── Convertir apoyos_global ──────────────────────────────────────────────
    apoyos_list = []
    for nome, d in apoyos_global.items():
        if es_placeholder(d.get('sn',0), d['m'], d['h'], d['total'], nome): continue
        apoyos_list.append({
            'apoyo':nome,'m':d['m'],'h':d['h'],'total':d['total'],
            'n_municipios':len(d['municipios']),
            'n_programas':len(d.get('programas',set())),
        })
    apoyos_list.sort(key=lambda x: -x['total'])

    # ── Convertir desglose: sort por total, sin filtro de mínimo ────────────
    desglose_final = {}
    for mun_dk, apoyos_dict in desglose.items():
        lista = sorted(
            [a for a in apoyos_dict.values()
             if not es_placeholder(a.get('sn',0), a['m'], a['h'], a['total'], a.get('apoyo',''))],
            key=lambda x: -x['total']
        )
        if lista: desglose_final[mun_dk] = lista

    return apoyos_list, desglose_final



# ─── SHEET: Beneficiarios Localizables ───────────────────────────────────────
def parse_localizables():
    """
    Lee la hoja 'Beneficiarios Localizables' y devuelve:
      - total, m, h  (grand total)
      - por_institucion: [{nombre, m, h, total,
                           rangos:{rango:n}, municipios:{mun:{m,h,total}}}]
      - por_municipio, m_por_municipio, h_por_municipio  (totales globales)
      - rangos_edad (total global)
    """
    if 'Beneficiarios Localizables' not in wb.sheetnames:
        return {'total':0,'m':0,'h':0,'por_institucion':[],'por_municipio':{},'rangos_edad':{}}

    rows = list(_WB_CACHE.get('Beneficiarios Localizables', []))
    INST_SET_LOC = {'DIF','ICHDII','ICHIJUV','SALUD','SDHYBC','SDHyBC','SDBYBC',
                    'SPYCI','SPyCI','CECYTECH','COESPO','COESVI','CULTURA','ICHD',
                    'ICHIMUJ','RURAL','SEECH','SEYD','SEyD','TRABAJO','TURISMO'}
    def _is_inst_l(s): return s.strip().upper() in {x.upper() for x in INST_SET_LOC}

    # ── Autodescubrimiento de columnas ancla ───────────────────────────────────
    # Busca en cada fila los textos clave que anclan cada tabla, sin asumir índices.
    #
    # col_a_data  : columna de la Tabla A  (texto "BENEFICIARIOS LOCALIZABLES")
    # col_b_data  : columna de la Tabla B  (texto "BENEFICIARIOS LOCALIZABLES POR RANGO DE EDAD")
    # col_c_data  : columna de la Tabla C  (texto "BENEFICIARIOS LOCALIZABLES POR MUNICIPIO")
    # hdr_row_idx : índice de la fila de encabezados (donde aparecen los tres textos ancla)

    col_a_data = col_b_data = col_c_data = None
    hdr_row_idx = None   # fila donde están los encabezados de Tabla B y Tabla C
    sub_hdr_a_row = None  # fila donde está el sub-encabezado de Tabla A (M | H | TOTAL)

    for ri, row in enumerate(rows):
        for ci, v in enumerate(row):
            vs = str(v or '').strip().upper()
            if vs == 'BENEFICIARIOS LOCALIZABLES' and col_a_data is None and ci < 5:
                col_a_data = ci
            if 'BENEFICIARIOS LOCALIZABLES POR RANGO DE EDAD' in vs and col_b_data is None:
                col_b_data = ci
                hdr_row_idx = ri   # esta fila tiene los rangos de edad Y los municipios de Tabla C
            if 'BENEFICIARIOS LOCALIZABLES POR MUNICIPIO' in vs and col_c_data is None:
                col_c_data = ci

    # Fallbacks
    if col_a_data is None: col_a_data = 0
    if col_b_data is None: col_b_data = 6
    if col_c_data is None: col_c_data = 17

    # ── Tabla A: INST → {m, h, total} ─────────────────────────────────────────
    # Encabezado de Tabla A: la fila inmediatamente siguiente al hdr_row_idx
    # contiene "BENEFICIARIOS LOCALIZABLES | M | H | TOTAL" en las sub-columnas.
    # Detectamos los índices de M, H, TOTAL dinámicamente.
    inst_totals = {}
    col_a_m = col_a_h = col_a_total = None
    total_a = m_tot = h_tot = 0

    # Encontrar la fila de sub-encabezado de Tabla A (contiene "M","H","TOTAL" en cols adyacentes)
    sub_hdr_a_idx = None
    for ri in range(hdr_row_idx if hdr_row_idx else 0, min(hdr_row_idx + 5 if hdr_row_idx else 10, len(rows))):
        row = rows[ri]
        v_a = str(row[col_a_data] or '').strip().upper()
        if v_a in ('BENEFICIARIOS LOCALIZABLES', 'BENEFICIARIOS LOCALIZABLES '):
            # Mapear sub-columnas: buscar M, H, TOTAL en las celdas a la derecha
            for ci in range(col_a_data + 1, min(col_a_data + 6, len(row))):
                h = str(row[ci] or '').strip().upper()
                if h == 'M'     and col_a_m     is None: col_a_m     = ci
                if h == 'H'     and col_a_h     is None: col_a_h     = ci
                if h == 'TOTAL' and col_a_total  is None: col_a_total = ci
            sub_hdr_a_idx = ri
            break

    if col_a_m     is None: col_a_m     = col_a_data + 1
    if col_a_h     is None: col_a_h     = col_a_data + 2
    if col_a_total is None: col_a_total = col_a_data + 3

    # Leer filas de datos de Tabla A hasta encontrar TOTAL
    reading_a = False
    for ri, row in enumerate(rows):
        if sub_hdr_a_idx is not None and ri <= sub_hdr_a_idx:
            reading_a = (ri == sub_hdr_a_idx)
            continue
        if not reading_a:
            continue
        nombre = str(row[col_a_data] or '').strip()
        if not nombre:
            continue
        if nombre.upper() == 'TOTAL':
            total_a = int(sf(row[col_a_total] if len(row) > col_a_total else None))
            m_tot   = int(sf(row[col_a_m]     if len(row) > col_a_m     else None))
            h_tot   = int(sf(row[col_a_h]     if len(row) > col_a_h     else None))
            break
        if _is_inst_l(nombre):
            inst_totals[nombre] = {
                'm':     int(sf(row[col_a_m]     if len(row) > col_a_m     else None)),
                'h':     int(sf(row[col_a_h]     if len(row) > col_a_h     else None)),
                'total': int(sf(row[col_a_total] if len(row) > col_a_total else None)),
            }

    total = total_a if total_a else sum(v['total'] for v in inst_totals.values())
    if not m_tot: m_tot = sum(v['m'] for v in inst_totals.values())
    if not h_tot: h_tot = sum(v['h'] for v in inst_totals.values())

    # ── Tabla B: INST → {rangos} — autodescubrimiento de sub-columnas ─────────
    # La fila hdr_row_idx contiene los encabezados de rangos a la derecha de col_b_data
    inst_rangos = {}
    col_b_rangos = {}   # {nombre_rango: col_index}
    if hdr_row_idx is not None:
        hdr_row = rows[hdr_row_idx]
        for ci in range(col_b_data + 1, len(hdr_row)):
            h = str(hdr_row[ci] or '').strip()
            if h and h.upper() != 'TOTAL':
                col_b_rangos[h] = ci
            elif h.upper() == 'TOTAL':
                col_b_total_idx = ci
                break

    for ri, row in enumerate(rows):
        if hdr_row_idx is not None and ri <= hdr_row_idx:
            continue
        g = str(row[col_b_data] or '').strip() if len(row) > col_b_data else ''
        if not g or g.upper() in ('TOTAL', 'M', 'H'):
            if g.upper() == 'TOTAL':
                # Leer rangos globales de esta fila TOTAL
                pass
            continue
        if _is_inst_l(g):
            inst_rangos[g] = {
                rng: int(sf(row[ci] if len(row) > ci else None))
                for rng, ci in col_b_rangos.items()
            }

    # ── Tabla B fila TOTAL: rangos_edad global ─────────────────────────────────
    rangos_edad = {}
    AGE_RANGO_NORM = {'0-5':'0-5','6-11':'6-11','12-17':'12-17','18-29':'18-29',
                      '30-49':'30-49','50-64':'50-64','65+':'65+',
                      'Sin datos':'sin_datos','SIN DATOS':'sin_datos','Sin dato':'sin_datos'}
    for ri, row in enumerate(rows):
        if hdr_row_idx is not None and ri <= hdr_row_idx:
            continue
        g = str(row[col_b_data] or '').strip() if len(row) > col_b_data else ''
        if g.upper() == 'TOTAL':
            for rng_label, ci in col_b_rangos.items():
                key_norm = AGE_RANGO_NORM.get(rng_label, rng_label)
                rangos_edad[key_norm] = int(sf(row[ci] if len(row) > ci else None))
            break

    # ── Tabla C: municipios — autodescubrimiento desde encabezado ─────────────
    # La fila hdr_row_idx contiene los nombres de municipio a la derecha de col_c_data.
    mun_cols_c = {}   # {col_index: nombre_municipio}
    if hdr_row_idx is not None:
        hdr_row = rows[hdr_row_idx]
        for ci in range(col_c_data + 1, len(hdr_row)):
            v = str(hdr_row[ci] or '').strip()
            if v and v.upper() != 'TOTAL':
                mun_cols_c[ci] = v
            elif v.upper() == 'TOTAL':
                break   # columna TOTAL al final — detenerse

    inst_muns  = {}   # {inst: {mun: {m, h, total}}}
    prog_muns  = {}   # {(inst, prog): {mun: {m, h, total}}}
    sexo_c     = None
    inst_c     = None
    prog_c     = None

    for ri, row in enumerate(rows):
        if hdr_row_idx is not None and ri <= hdr_row_idx:
            continue
        r_lbl = str(row[col_c_data] or '').strip() if len(row) > col_c_data else ''
        if not r_lbl or 'BENEFICIARIOS LOCALIZABLES POR MUNICIPIO' in r_lbl.upper():
            continue
        if r_lbl.upper() == 'M':
            sexo_c = 'M'; prog_c = None; continue
        if r_lbl.upper() == 'H':
            sexo_c = 'H'; prog_c = None; continue
        if r_lbl.upper() == 'TOTAL':
            prog_c = None; continue
        if _is_inst_l(r_lbl):
            inst_c = r_lbl; prog_c = None
            inst_muns.setdefault(r_lbl, {})
            for ci, mun in mun_cols_c.items():
                if mun not in inst_muns[r_lbl]:
                    inst_muns[r_lbl][mun] = {'m': 0.0, 'h': 0.0}
                val = sf(row[ci] if len(row) > ci else None)
                if sexo_c == 'M':   inst_muns[r_lbl][mun]['m'] = val
                elif sexo_c == 'H': inst_muns[r_lbl][mun]['h'] = val
        else:
            if inst_c is not None:
                prog_c = r_lbl
                key = (inst_c, prog_c)
                prog_muns.setdefault(key, {})
                for ci, mun in mun_cols_c.items():
                    if mun not in prog_muns[key]:
                        prog_muns[key][mun] = {'m': 0.0, 'h': 0.0}
                    val = sf(row[ci] if len(row) > ci else None)
                    if sexo_c == 'M':   prog_muns[key][mun]['m'] = val
                    elif sexo_c == 'H': prog_muns[key][mun]['h'] = val

    for inst in inst_muns:
        for mun in inst_muns[inst]:
            d = inst_muns[inst][mun]
            d['total'] = d['m'] + d['h']
    for key in prog_muns:
        for mun in prog_muns[key]:
            d = prog_muns[key][mun]
            d['total'] = d['m'] + d['h']

    # ── por_institucion combinado ──────────────────────────────────────────────
    por_institucion = []
    for nombre, totals in inst_totals.items():
        por_institucion.append({
            'nombre':     nombre,
            'm':          totals['m'],
            'h':          totals['h'],
            'total':      totals['total'],
            'rangos':     inst_rangos.get(nombre, {}),
            'municipios': inst_muns.get(nombre, {}),
        })

    # ── por_municipio global: filas M / H / TOTAL en Tabla C ──────────────────
    # Busca las primeras filas donde col_c_data = 'M', 'H', 'TOTAL' (globales)
    row_m_mun = row_h_mun = row_tot = []
    found_m = found_h = found_tot = False
    for ri, row in enumerate(rows):
        if hdr_row_idx is not None and ri <= hdr_row_idx:
            continue
        r_lbl = str(row[col_c_data] or '').strip().upper() if len(row) > col_c_data else ''
        if r_lbl == 'M'     and not found_m:   row_m_mun = row; found_m   = True
        if r_lbl == 'H'     and not found_h:   row_h_mun = row; found_h   = True
        if r_lbl == 'TOTAL' and not found_tot: row_tot   = row; found_tot = True
        if found_m and found_h and found_tot:
            break

    por_municipio = {}; m_por_municipio = {}; h_por_municipio = {}
    for ci, mun_name in mun_cols_c.items():
        key  = _norm(str(mun_name).strip())
        vtot = int(sf(row_tot[ci]   if ci < len(row_tot)   else 0))
        vm   = int(sf(row_m_mun[ci] if ci < len(row_m_mun) else 0))
        vh   = int(sf(row_h_mun[ci] if ci < len(row_h_mun) else 0))
        if vtot > 0:
            por_municipio[key]   = vtot
            m_por_municipio[key] = vm
            h_por_municipio[key] = vh

    # Mapa de aliases: variaciones conocidas en nombres del Excel → nombre canónico normalizado
    # Formato: {NORM_INST||NORM_PROG_EXCEL: NORM_INST||NORM_PROG_CANONICAL}
    PROG_ALIASES = {
        'SDHYBC||APOYOS A PERSONAS A DISCAPACIDAD':
            'SDHYBC||APOYOS A PERSONAS CON DISCAPACIDAD',
    }

    # Exportar prog_muns con claves normalizadas (sin tildes) para cruce robusto
    # Clave: "NORM_INST||NORM_PROG" -> {NORM_MUN: {m, h, total}}
    prog_muns_norm = {}
    for (inst_k, prog_k), mun_dict in prog_muns.items():
        nkey = f'{_norm(inst_k)}||{_norm(prog_k)}'
        nkey = PROG_ALIASES.get(nkey, nkey)   # aplicar alias si existe
        norm_mun_dict = {_norm(mk): mv for mk, mv in mun_dict.items()}
        if nkey in prog_muns_norm:
            # Fusionar si el alias ya existe (sumar m, h, total por municipio)
            for mk, mv in norm_mun_dict.items():
                if mk in prog_muns_norm[nkey]:
                    prog_muns_norm[nkey][mk]['m']     += mv['m']
                    prog_muns_norm[nkey][mk]['h']     += mv['h']
                    prog_muns_norm[nkey][mk]['total'] += mv['total']
                else:
                    prog_muns_norm[nkey][mk] = dict(mv)
        else:
            prog_muns_norm[nkey] = norm_mun_dict

    return {
        'total':           total,
        'm':               m_tot,
        'h':               h_tot,
        'por_institucion': por_institucion,
        'por_municipio':   por_municipio,
        'm_por_municipio': m_por_municipio,
        'h_por_municipio': h_por_municipio,
        'rangos_edad':     rangos_edad,
        'prog_muns':       prog_muns_norm,
    }

# ─── SHEET: Indicadores y Metas ──────────────────────────────────────────────
def parse_indicadores():
    """
    Lee la hoja 'Indicadores y Metas' y devuelve lista de programas.
    Las columnas se descubren dinámicamente desde la fila de encabezado,
    por lo que son robustas a reordenamientos o inserciones de columnas en Power Query.

    Encabezados buscados (case-insensitive, sin tildes):
      Institución, Programa Social, Población Potencial, Población Objetivo,
      Población Alcanzada, Poblacion Única, Beneficiarios Reales, Mujeres,
      Hombres, Sin identificar, Presupuesto, Gasto, EP,
      Metas programadas, Avance metas, Eficacia %, Eficiencia %, Desempeño %
    Las columnas A (clave) y B (nombre) se detectan como las primeras dos
    columnas con datos en la fila de encabezado.
    """
    if 'Indicadores y Metas' not in wb.sheetnames:
        return []

    rows = list(_WB_CACHE.get('Indicadores y Metas', []))

    def parse_pct(v):
        if v is None: return None
        if isinstance(v, (int, float)): return float(v)
        s = str(v).strip()
        if s in ('', '#DIV/0!', 'No es posible calcular', '#REF!'): return None
        s = s.replace('%','').strip()
        try:
            f = float(s)
            return f if f <= 1.0 else f / 100.0
        except: return None

    def parse_num(v):
        if v is None: return None
        if isinstance(v, (int, float)): return float(v) if float(v) > 0 else None
        s = str(v).strip()
        if s in ('', '#DIV/0!', 'No es posible calcular', '#REF!'): return None
        try: return float(s) if float(s) > 0 else None
        except: return None

    # ── Autodescubrimiento de columnas desde la fila de encabezado ────────────
    # Mapa: texto_encabezado_normalizado → nombre_campo_interno
    _HDR_MAP = {
        'institucion':                'col_inst',
        'programa social':            'col_prog_social',
        'poblacion potencial':        'col_pob_pot',
        'poblacion objetivo':         'col_pob_obj',
        'poblacion alcanzada':        'col_pob_alc',
        'poblacion unica':            'col_benef_unicos',
        'beneficiarios unicos':       'col_benef_unicos',
        'beneficiarios reales':       'col_benef_reales',
        'mujeres':                    'col_mujeres',
        'hombres':                    'col_hombres',
        'sin identificar':            'col_sin_id',
        'presupuesto':                'col_pres',
        'gasto':                      'col_gasto',
        'ep':                         'col_ep',
        'metas programadas':          'col_metas',
        'avance metas':               'col_avance',
        'eficacia %':                 'col_eficacia',
        'eficiencia %':               'col_eficiencia',
        'desempeno %':                'col_desempeno',
    }

    def _nhdr(s):
        """Normaliza encabezado: minúsculas, sin tildes, sin espacios extra."""
        import unicodedata
        s = unicodedata.normalize('NFD', str(s or '').lower()).encode('ascii', 'ignore').decode()
        return ' '.join(s.split())

    # Buscar la fila de encabezados (contiene "Institución" o "Programa Social")
    hdr_row_idx_ind = None
    col_map = {}   # {campo: col_index}
    col_clave = col_nombre = None

    for ri, row in enumerate(rows):
        non_null = [(ci, v) for ci, v in enumerate(row) if v is not None]
        if not non_null:
            continue
        # Detectar si esta fila es la de encabezados
        has_inst = any('instituci' in _nhdr(str(v)) for _, v in non_null)
        has_prog = any('programa social' in _nhdr(str(v)) for _, v in non_null)
        if has_inst or has_prog:
            hdr_row_idx_ind = ri
            # Mapear cada columna
            for ci, v in non_null:
                norm = _nhdr(str(v))
                field = _HDR_MAP.get(norm)
                if field:
                    col_map[field] = ci
            # Las dos primeras columnas con datos son clave y nombre
            cols_with_data = [ci for ci, v in non_null]
            first_two = sorted(c for c in cols_with_data if c not in col_map.values())[:2]
            if len(first_two) >= 2:
                col_clave, col_nombre = first_two[0], first_two[1]
            elif len(first_two) == 1:
                col_clave = first_two[0]
            break

    # Posiciones FIJAS para esta hoja (A=clave, B=nombre, C=institución)
    col_clave  = 0
    col_nombre = 1
    col_inst   = col_map.get('col_inst', 2)
    col_pob_pot   = col_map.get('col_pob_pot',      4)
    col_pob_obj   = col_map.get('col_pob_obj',      5)
    col_pob_alc   = col_map.get('col_pob_alc',      6)
    col_benef_uni = col_map.get('col_benef_unicos',  8)  # col I = Beneficiarios Únicos
    col_benef_rea = col_map.get('col_benef_reales',  7)  # col H = Beneficiarios Reales
    col_mujeres   = col_map.get('col_mujeres',       9)
    col_hombres   = col_map.get('col_hombres',      10)
    col_sin_id    = col_map.get('col_sin_id',        11)
    col_pres      = col_map.get('col_pres',          12)
    col_gasto     = col_map.get('col_gasto',         13)
    col_ep        = col_map.get('col_ep',            14)
    col_metas     = col_map.get('col_metas',         15)
    col_avance    = col_map.get('col_avance',        16)
    col_eficacia  = col_map.get('col_eficacia',      18)
    col_eficiencia= col_map.get('col_eficiencia',    19)
    col_desempeno = col_map.get('col_desempeno',     20)

    def gc(row, ci):
        """Obtiene celda de forma segura."""
        return row[ci] if ci is not None and ci < len(row) else None

    programas = []
    inst_act = None
    data_start = (hdr_row_idx_ind + 1) if hdr_row_idx_ind is not None else 1
    for r in rows[data_start:]:
        nombre = str(gc(r, col_nombre) or '').strip()
        if not nombre:
            continue
        inst_col = str(gc(r, col_inst) or '').strip()
        if inst_col:
            inst_act = inst_col
        clave = str(gc(r, col_clave) or '').strip() or 'N/A'

        prog = {
            'institucion':   inst_act or '',
            'clave':         clave,
            'nombre':        nombre,
            'pob_potencial': parse_num(gc(r, col_pob_pot)),
            'pob_objetivo':  parse_num(gc(r, col_pob_obj)),
            'pob_alcanzada': parse_num(gc(r, col_pob_alc)),
            'benef_unicos':  parse_num(gc(r, col_benef_uni)),
            'benef_reales':  parse_num(gc(r, col_benef_rea)),
            'mujeres':       parse_num(gc(r, col_mujeres)),
            'hombres':       parse_num(gc(r, col_hombres)),
            'sin_id':        parse_num(gc(r, col_sin_id)),
            'presupuesto':   parse_num(gc(r, col_pres)),
            'gasto':         parse_num(gc(r, col_gasto)),
            'ep':            parse_pct(gc(r, col_ep)),
            'metas_prog':    parse_num(gc(r, col_metas)),
            'avance_metas':  parse_num(gc(r, col_avance)),
            'eficacia':      parse_pct(gc(r, col_eficacia)),
            'eficiencia':    parse_pct(gc(r, col_eficiencia)),
            'desempeño':     parse_pct(gc(r, col_desempeno)),
        }
        programas.append(prog)

    return programas


# ─── MAIN ─────────────────────────────────────────────────────────────────────
# ─── SHEET 3: Tabla G3 — Apoyos por Dependencia/Programa ────────────────────
def parse_apoyos_g3():
    """
    Lee la tabla G3 de Apoyos Otorgados: Dependencia → Programa, columnas SN/M/H/TOTAL.
    Devuelve {inst_key: {sn,m,h,total, programas:{prog:{sn,m,h,total}}}}

    Autodescubrimiento: localiza la tabla buscando el texto "APOYOS POR PROGRAMA"
    en cualquier celda, luego mapea columnas SN/M/H/TOTAL desde esa fila de encabezado.
    Esto hace la función resistente a que Power Query agregue columnas o mueva la tabla.
    """
    rows3 = list(_WB_CACHE.get('Apoyos Otorgados', []))

    # ── 1. Localizar fila de encabezado y columna ancla de la tabla G3 ────────
    # El encabezado contiene "APOYOS POR PROGRAMA" en alguna celda.
    # La columna de esa celda es col_g_data (columna de nombres de inst/prog).
    col_g_data = None
    col_g_sn = col_g_m = col_g_h = col_g_total = None
    data_start_g3 = 0

    for ri, row in enumerate(rows3):
        for ci, v in enumerate(row):
            if v and 'APOYOS POR PROGRAMA' in str(v).upper():
                col_g_data = ci
                # Mapear sub-columnas SN/M/H/TOTAL desde esta misma fila
                for cj in range(ci + 1, len(row)):
                    h = str(row[cj] or '').strip().upper()
                    if h in ('SIN DATOS', 'SNDATOS', 'SIN DATO', 'SN') and col_g_sn is None:
                        col_g_sn = cj
                    elif h == 'M'     and col_g_m     is None: col_g_m     = cj
                    elif h == 'H'     and col_g_h     is None: col_g_h     = cj
                    elif h == 'TOTAL' and col_g_total  is None: col_g_total = cj
                data_start_g3 = ri + 1
                break
        if col_g_data is not None:
            break

    # Fallbacks a posiciones conocidas si el encabezado no se encuentra
    if col_g_data  is None: col_g_data  = 6
    if col_g_sn    is None: col_g_sn    = col_g_data + 1
    if col_g_m     is None: col_g_m     = col_g_data + 2
    if col_g_h     is None: col_g_h     = col_g_data + 3
    if col_g_total is None: col_g_total = col_g_data + 4

    def _gc3(row, ci):
        return row[ci] if ci is not None and ci < len(row) else None

    # ── 2. Leer datos ──────────────────────────────────────────────────────────
    inst_act = None
    result = {}
    for row in rows3[data_start_g3:]:
        g = clean(_gc3(row, col_g_data))
        if not g: continue
        if g.upper() == 'TOTAL': continue
        sn3  = sf(_gc3(row, col_g_sn))
        m3   = sf(_gc3(row, col_g_m))
        h3   = sf(_gc3(row, col_g_h))
        tot3 = sf(_gc3(row, col_g_total))
        if is_inst(g):
            inst_act = g
            result[inst_act] = {'sn': sn3, 'm': m3, 'h': h3, 'total': tot3, 'programas': {}}
        elif inst_act and tot3 > 0:
            result[inst_act]['programas'][corr(g)] = {'sn': sn3, 'm': m3, 'h': h3, 'total': tot3}
    return result


# ─── SHEET 3b: Tabla A3 — Apoyos por APOYO/Dependencia/Programa ─────────────
def parse_apoyos_a3():
    """
    Lee tabla A3: APOYO → Dependencia → Programa → Municipio → Edad
    Devuelve {inst: {apoyo: {prog: {sn,m,h,total}}}}
    Usa lookahead: una fila es APOYO si la siguiente fila no-edad es una INST.

    Autodescubrimiento: localiza el inicio de datos buscando el encabezado
    "TIPOS DE APOYOS" en la columna A (o el texto ancla equivalente), en lugar
    de asumir que siempre son exactamente 4 filas de encabezado.
    """
    INST_SET_U = {x.upper() for x in
                  ['CECYTECH','COESPO','COESVI','CULTURA','DIF','ICHD','ICHDII','ICHIJUV',
                   'ICHIMUJ','RURAL','SALUD','SDHYBC','SDHyBC','SEECH','SEYD','SEyD',
                   'SPYCI','SPyCI','TRABAJO','TURISMO']}
    SKIP_U = {'TOTAL','NUMERO DE BENEFICIARIOS UNICOS POR MUNIC',
              'NUMERO DE BENEFICIARIOS UNICOS POR MUNICIPIO','RECUENTO DE # DE APOYOS'}
    AGES_U = {'0-5','6-11','12-17','18-29','30-49','50-64','65+'}
    # Textos que indican fila de encabezado (no datos) en la Tabla A3
    HDR_MARKERS = {'TIPOS DE APOYOS', 'RECUENTO DE # DE APOYOS',
                   'MES CORRESPONDIENTE', 'ALL', 'BENEFICIARIOS UNICOS POR PROGRAMA'}

    def _is_inst(s): return s.strip().upper() in INST_SET_U
    def _is_age(s):
        u = s.strip().upper()
        return u in AGES_U or 'SIN DATO' in u

    rows3 = list(_WB_CACHE.get('Apoyos Otorgados', []))

    # ── 1. Encontrar el inicio real de datos en la Tabla A3 ───────────────────
    # La tabla A3 comienza en la primera fila donde la col A contiene un valor
    # que NO es un encabezado / marcador de slicer y NO es una fila de rango de edad.
    # En la práctica es la primera fila con un nombre de apoyo o institución.
    data_start_a3 = 0
    for ri, row in enumerate(rows3):
        v = str(row[0] or '').strip()
        if not v:
            continue
        vu = v.upper()
        if vu in HDR_MARKERS or vu in SKIP_U or _is_age(v):
            continue
        # Primera fila con contenido real
        data_start_a3 = ri
        break

    # ── 2. Filtrar filas relevantes desde el inicio detectado ─────────────────
    clean_rows = []
    for row in rows3[data_start_a3:]:
        val = str(row[0] or '').strip()
        if not val or val.upper() in SKIP_U: continue
        if _is_age(val): continue
        clean_rows.append((val, sf(row[1] if len(row)>1 else None),
                               sf(row[2] if len(row)>2 else None),
                               sf(row[3] if len(row)>3 else None),
                               sf(row[4] if len(row)>4 else None)))

    # ── 3. Lookahead y construcción del resultado ─────────────────────────────
    next_inst = [False] * len(clean_rows)
    for idx in range(len(clean_rows)-1):
        next_inst[idx] = _is_inst(clean_rows[idx+1][0])

    apoyo_act = inst_act = prog_act = None
    result = {}
    for idx, (val, sn3, m3, h3, t3) in enumerate(clean_rows):
        if _is_inst(val):
            inst_act = val; prog_act = None; continue
        if inst_act is None or next_inst[idx]:
            apoyo_act = corr(val); inst_act = None; prog_act = None; continue
        if prog_act is None:
            prog_act = corr(val)
            result.setdefault(inst_act, {}).setdefault(apoyo_act, {})[prog_act] = {
                'sn': sn3, 'm': m3, 'h': h3, 'total': t3
            }
            continue
        # Municipio / edad — ignorar
    return result


# ─── MAPA DE COLUMNAS DE RANGOS (Tabla B hoja 1 y hoja Localizables) ─────────
# Fila de encabezado (fila 4): col G = label, H=0-5, I=6-11, J=12-17, K=18-29,
#                                        L=30-49, M=50-64, N=65+, O=Sin datos, P=TOTAL
_RANGO_COL = {'0-5': 7, '6-11': 8, '12-17': 9, '18-29': 10,
              '30-49': 11, '50-64': 12, '65+': 13, 'sin_datos': 14}

def _col_indices(rangos_edad):
    """Devuelve lista de índices de columna para los rangos solicitados."""
    if not rangos_edad:
        return list(_RANGO_COL.values())
    return [_RANGO_COL[r] for r in rangos_edad if r in _RANGO_COL]


# ─── NUEVA TABLA "Apoyos Otorgados" — cols T en adelante ─────────────────────
# Estructura: FILAS = RANGO_EDAD → INST → PROGRAMA → APOYO
#             COLS  = bloque M (cols 20-91) + bloque H (cols 93-161)
#                     col 92  = Total M global
#                     col 162 = Total H global  (FG)
#                     col 163 = Total M+H+sn    (FH)
# Fila 6 (índ. 5) = encabezado de municipios en ambos bloques.
# Filas de RANGO: col 19 contiene '0-5','6-11','12-17','18-29','30-49','50-64','65+'
# Bajo cada rango: INST → PROGRAMA → APOYO (todos en col 19)

def _build_nueva_tabla_mapa():
    """Construye el mapa municipio→(col_m, col_h) desde la fila de encabezado."""
    rows3 = list(_WB_CACHE.get('Apoyos Otorgados', []))[5:6]
    if not rows3: return {}, {}
    row6 = rows3[0]
    SKIP = {'M','H','SIN DATOS','TOTAL SIN DATOS','TOTAL M',''}
    mun_m = {}  # {MUN_NORM: col_idx}
    mun_h = {}
    for j in range(20, 92):   # bloque mujeres
        v = str(row6[j] if j < len(row6) and row6[j] else '').strip().upper()
        if v and v not in SKIP:
            mun_m[_norm(v)] = j
    for j in range(93, 162):  # bloque hombres
        v = str(row6[j] if j < len(row6) and row6[j] else '').strip().upper()
        if v and v not in SKIP:
            mun_h[_norm(v)] = j
    return mun_m, mun_h

_NT_MUN_M, _NT_MUN_H = _build_nueva_tabla_mapa()
_NT_COL_TOT   = 163   # FH = Total M+H+sn
_NT_COL_TOT_H = 162   # FG = Total H
_NT_COL_TOT_M = 92    # Total M

_NT_AGE_KEYS = {'0-5','6-11','12-17','18-29','30-49','50-64','65+'}
_NT_INST_U   = {x.upper() for x in
                ['CECYTECH','COESPO','COESVI','CULTURA','DIF','ICHD','ICHDII','ICHIJUV',
                 'ICHIMUJ','RURAL','SALUD','SDHYBC','SDBYBC','SEECH','SEYD','SEYD',
                 'SPYCI','SPYCI','TRABAJO','TURISMO']}

def parse_apoyos_nueva_tabla(inst_key, rangos_edad=None, sexo=None):
    """
    Lee la nueva tabla (cols T+ de 'Apoyos Otorgados').
    Filas: RANGO_EDAD → INST → PROGRAMA → APOYO
    Cols:  bloque M (cols 20-91) + bloque H (cols 93-161) + totales (cols 92,162,163)

    Retorna:
      total, m, h
      por_programa: [{nombre, total, m, h}]  — apoyos por programa de la inst
      por_municipio: [{municipio, total, m, h}]  — apoyos por municipio
    """
    inst_norm    = _norm(inst_key) if inst_key else None
    rangos_set   = set(rangos_edad) if rangos_edad else None

    rows3 = list(_WB_CACHE.get('Apoyos Otorgados', []))

    total_t = m_t = h_t = 0.0
    por_programa  = {}  # {nombre: {total, m, h}}
    por_municipio = {}  # {mun_norm: {nombre, total, m, h}}

    cur_age  = None
    cur_inst = None   # None si no es la inst que buscamos
    in_inst  = False

    for row in rows3:
        v19 = str(row[19] if len(row) > 19 and row[19] else '').strip()
        if not v19: continue

        v19u = v19.upper()

        # ── Rango de edad ──────────────────────────────────────────────────────
        if v19u in _NT_AGE_KEYS:
            cur_age  = v19u
            cur_inst = None
            in_inst  = False
            continue

        if cur_age is None: continue

        # ── Filtro de rangos ───────────────────────────────────────────────────
        if rangos_set and cur_age not in rangos_set:
            continue

        # ── Institución ────────────────────────────────────────────────────────
        if v19u in _NT_INST_U:
            if inst_norm is None or _norm(v19) == inst_norm:
                in_inst  = True
                cur_inst = v19
                # Totales desde la fila de institución (fuente canónica, sin doble conteo)
                t = sf(row[_NT_COL_TOT]   if _NT_COL_TOT   < len(row) else None)
                m = sf(row[_NT_COL_TOT_M] if _NT_COL_TOT_M < len(row) else None)
                h = sf(row[_NT_COL_TOT_H] if _NT_COL_TOT_H < len(row) else None)
                total_t += t
                m_t     += m
                h_t     += h
                # Apoyos por municipio desde la fila de institución
                for mun_n, col_m in _NT_MUN_M.items():
                    vm = sf(row[col_m] if col_m < len(row) else None)
                    col_h = _NT_MUN_H.get(mun_n)
                    vh = sf(row[col_h] if col_h and col_h < len(row) else None)
                    if vm + vh > 0:
                        if mun_n not in por_municipio:
                            por_municipio[mun_n] = {'nombre': v19u, 'total': 0.0, 'm': 0.0, 'h': 0.0}
                        por_municipio[mun_n]['m']     += vm
                        por_municipio[mun_n]['h']     += vh
                        por_municipio[mun_n]['total'] += vm + vh
            else:
                # Es otra institución — dejar de acumular sub-filas
                in_inst  = False
                cur_inst = None
            continue

        if not in_inst: continue

        # ── Sub-fila bajo la institución buscada (programa/apoyo) ─────────────
        # Solo acumular si la fila pertenece a nuestra institución (in_inst=True)
        # Los totales globales ya están en la fila de INST; aquí solo llenamos por_programa
        t = sf(row[_NT_COL_TOT]   if _NT_COL_TOT   < len(row) else None)
        m = sf(row[_NT_COL_TOT_M] if _NT_COL_TOT_M < len(row) else None)
        h = sf(row[_NT_COL_TOT_H] if _NT_COL_TOT_H < len(row) else None)
        prog_key = corr(v19)
        if prog_key not in por_programa:
            por_programa[prog_key] = {'total': 0.0, 'm': 0.0, 'h': 0.0}
        por_programa[prog_key]['total'] += t
        por_programa[prog_key]['m']     += m
        por_programa[prog_key]['h']     += h

    # Aplicar filtro de sexo al total
    if   sexo == 'm': total_t = m_t
    elif sexo == 'h': total_t = h_t

    # Buscar nombre real de municipio (sin normalizar)
    row6 = rows3[5] if len(rows3) > 5 else []
    def _nombre_mun(mun_n):
        col_m = _NT_MUN_M.get(mun_n)
        if col_m and col_m < len(row6) and row6[col_m]:
            return str(row6[col_m]).strip()
        return mun_n

    return {
        'total': total_t,
        'm':     m_t,
        'h':     h_t,
        'por_programa':  sorted(
            [{'nombre': k, **v} for k, v in por_programa.items() if v['total'] > 0],
            key=lambda x: -x['total']
        ),
        'por_municipio': sorted(
            [{'municipio': _nombre_mun(k), **{kk: vv for kk,vv in v.items() if kk != 'nombre'}}
             for k, v in por_municipio.items() if v['total'] > 0],
            key=lambda x: -x['total']
        ),
    }


# ─── FUNCIÓN DE FILTROS CRUZADOS ─────────────────────────────────────────────
def calcular_filtro(instituciones_data, municipios_data, inst_key=None, mun_key=None,
                    sexo=None, rangos_edad=None):
    """
    Calcula beneficiarios filtrados leyendo DIRECTAMENTE de las hojas del Excel.
    Fuentes exactas según mapeo validado manualmente:

    BENEFICIARIOS ÚNICOS     → Hoja "Únicos y Rango de Edad", Tabla B
                               Fila de inst + cols de rangos (total)
                               Filas M/H de cada programa + cols de rangos (por sexo)
    APOYOS ENTREGADOS        → Hoja "Apoyos Otorgados", Tabla A
                               Filas de rango bajo municipios del inst filtrado
    BENEF. POR MUNICIPIO     → Hoja "Beneficiarios por Municipio", Tabla A
                               Filas de rango bajo programas del inst filtrado
    LOCALIZABLES             → Hoja "Beneficiarios Localizables", Tabla B
                               Fila de inst + filas H/M + cols de rangos
    LOCALIZABLES POR MUN     → Hoja "Beneficiarios Localizables", Tabla C
                               Filas de rango bajo programas del inst + col del municipio

    Parámetros:
      inst_key    : sigla de institución (ej. 'SDHyBC') — None = todas
      mun_key     : nombre de municipio  (ej. 'ALDAMA') — None = todos
      sexo        : 'm', 'h' o None (= m+h+sn)
      rangos_edad : lista de claves (ej. ['0-5','6-11','12-17','18-29']) — None = todos
    """
    cols = _col_indices(rangos_edad)
    inst_norm = _norm(inst_key) if inst_key else None
    mun_norm  = _norm(mun_key)  if mun_key  else None
    AGE_KEYS  = set(rangos_edad) if rangos_edad else set(_RANGO_COL.keys())

    def _sum_cols(row, col_list):
        return sum(sf(row[c] if c < len(row) else None) for c in col_list)

    # ── 1. BENEFICIARIOS ÚNICOS (Tabla B hoja "Únicos y Rango de Edad") ───────
    rows1 = list(_WB_CACHE.get('Unicos y Rango de Edad', []))

    total_b = m_b = h_b = 0.0
    por_rango_b = {r: {'m': 0.0, 'h': 0.0, 'total': 0.0} for r in (rangos_edad or list(_RANGO_COL.keys()))}
    por_programa_b = []
    inst_act1 = None
    prog_act1 = None
    ultimo_prog_idx = None

    # Guardar total directo de fila de institución (incluye sn)
    total_inst_row = 0.0

    for ri, row in enumerate(rows1):
        ng = clean(row[6] if len(row) > 6 else None)
        if not ng: continue
        ng_u = ng.strip().upper()

        if ng_u in ('TOTAL', 'MES CORRESPONDIENTE', 'BENEFICIARIOS ÚNICOS', 'BENEFICIARIOS UNICOS'):
            continue
        if 'BENEFICIARIOS UNICOS POR RANGO DE EDAD' in ng_u:
            continue

        if is_inst(ng):
            if inst_norm is None or _norm(ng) == inst_norm:
                inst_act1 = ng
                # Fila de institución: total por rango (incluye sin_dato_sexo)
                if inst_norm is not None:
                    total_inst_row = _sum_cols(row, cols)
            else:
                inst_act1 = None
            prog_act1 = None
            continue

        if inst_act1 is None:
            continue

        # ¿Es fila de sexo (M/H)?
        if ng_u in ('M', 'H'):
            if prog_act1 is not None and ultimo_prog_idx is not None:
                # Acumular M o H para el programa actual
                vm = _sum_cols(row, cols)
                if ng_u == 'M':
                    m_b += vm
                    if ultimo_prog_idx < len(por_programa_b):
                        por_programa_b[ultimo_prog_idx]['m'] += vm
                        for r_k, c_k in _RANGO_COL.items():
                            if c_k in cols and r_k in por_rango_b:
                                por_rango_b[r_k]['m'] += sf(row[c_k] if c_k < len(row) else None)
                else:
                    h_b += vm
                    if ultimo_prog_idx < len(por_programa_b):
                        por_programa_b[ultimo_prog_idx]['h'] += vm
                        for r_k, c_k in _RANGO_COL.items():
                            if c_k in cols and r_k in por_rango_b:
                                por_rango_b[r_k]['h'] += sf(row[c_k] if c_k < len(row) else None)
            continue

        if is_age(ng):
            continue

        # Es fila de programa
        prog_act1 = corr(ng)
        prog_total = _sum_cols(row, cols)
        if prog_total > 0 or True:
            por_programa_b.append({'nombre': prog_act1, 'institucion': inst_act1,
                                   'm': 0.0, 'h': 0.0, 'total': prog_total})
            ultimo_prog_idx = len(por_programa_b) - 1
            # Acumular total por rango
            for r_k, c_k in _RANGO_COL.items():
                if c_k in cols and r_k in por_rango_b:
                    por_rango_b[r_k]['total'] += sf(row[c_k] if c_k < len(row) else None)

    # Completar total por rango
    for r_k in por_rango_b:
        por_rango_b[r_k]['total'] = por_rango_b[r_k]['m'] + por_rango_b[r_k]['h']

    # Aplicar filtro de sexo al total global
    # - sexo=None → usar total de la fila de institución (incluye sn, es el número correcto)
    # - sexo='m'  → usar suma de filas M
    # - sexo='h'  → usar suma de filas H
    if sexo == 'm':
        total_b = m_b
    elif sexo == 'h':
        total_b = h_b
    else:
        total_b = total_inst_row  # fuente canónica: fila de inst en Tabla B

    # ── 2. APOYOS ENTREGADOS — Nueva tabla (cols T+) ────────────────────────────
    # Fuente canónica para totales y por_municipio: parse_apoyos_nueva_tabla()
    nt = parse_apoyos_nueva_tabla(
        inst_key   = inst_key,
        rangos_edad= list(AGE_KEYS) if rangos_edad else None,
        sexo       = sexo,
    )
    total_apoyos_f = sf(nt.get('total', 0))
    m_apoyos_f     = sf(nt.get('m', 0))
    h_apoyos_f     = sf(nt.get('h', 0))

    # Por municipio desde nueva tabla
    por_municipio_f = {}
    for entry in nt.get('por_municipio', []):
        mun_n = _norm(entry['municipio'])
        por_municipio_f[mun_n] = {
            'municipio': entry['municipio'],
            'm':         entry.get('m', 0),
            'h':         entry.get('h', 0),
            'total':     entry.get('total', 0),
        }

    # Por programa desde nueva tabla (para tabla 1 del reporte)
    apoyos_por_programa_nt = {_norm(e['nombre']): e for e in nt.get('por_programa', [])}

    # ── 2b. APOYOS POR TIPO (Tabla A izquierda) con pre-tagging ─────────────────
    # Fuente para sección 4: Tabla A izquierda — APOYO → INST → PROG → MUN → EDAD
    # Usa el mismo pre-tagging que parse_sheet3_full para clasificar correctamente.
    rows3_left = list(_WB_CACHE.get('Apoyos Otorgados', []))

    apoyos_por_tipo = {}   # {apoyo: {m, h, total}}

    _SKIP_L = {'RECUENTO DE # DE APOYOS', 'NUMERO DE BENEFICIARIOS UNICOS POR MUNICIPIO',
               'NUMERO DE BENEFICIARIOS UNICOS POR MUNIC', 'TOTAL',
               'TIPOS DE APOYOS', 'MES CORRESPONDIENTE', ''}
    def _next_real_l(i):
        for j in range(i + 1, len(rows3_left)):
            v = str(rows3_left[j][0] or '').strip()
            if not v or v.upper() in _SKIP_L: continue
            if is_age(v): continue
            return v
        return None

    # Pre-tagging idéntico a parse_sheet3_full
    tagged_l = []
    for i, row in enumerate(rows3_left):
        n = clean(row[0] if len(row) > 0 else None)
        if not n or n.upper() in _SKIP_L:
            tagged_l.append(('SKIP', n, row)); continue
        if is_age(n):
            tagged_l.append(('AGE', n, row)); continue
        if is_inst(n):
            tagged_l.append(('INST', n, row)); continue
        nr = _next_real_l(i)
        if nr and is_inst(nr):
            n_u = n.strip().upper()
            ESPECIALES_AP = {'NO IDENTIFICADO', 'FORANEO', 'FORÁNEO'}
            if n_u in ESPECIALES_AP or _norm(n) not in _MUN_VALIDOS_NORM:
                tagged_l.append(('APOYO', n, row)); continue
        tagged_l.append(('OTHER', n, row))

    apoyo_l = inst_l = prog_l = None
    en_inst_l = False

    for tag, name, row in tagged_l:
        if tag == 'SKIP': continue
        if tag == 'APOYO':
            apoyo_l = corr(name); inst_l = None; en_inst_l = False; prog_l = None
            continue
        if tag == 'INST':
            if inst_norm is None or _norm(name) == inst_norm:
                inst_l = name; en_inst_l = True
            else:
                inst_l = None; en_inst_l = False
            prog_l = None
            continue
        if tag == 'AGE':
            ak = norm_age(name)
            if ak and ak in AGE_KEYS and en_inst_l and apoyo_l:
                vm = sf(row[2] if len(row) > 2 else None)
                vh = sf(row[3] if len(row) > 3 else None)
                vt = sf(row[4] if len(row) > 4 else None) or vm + vh
                if sexo == 'm':    contrib = vm
                elif sexo == 'h':  contrib = vh
                else:              contrib = vt
                if apoyo_l not in apoyos_por_tipo:
                    apoyos_por_tipo[apoyo_l] = {'m': 0.0, 'h': 0.0, 'total': 0.0}
                apoyos_por_tipo[apoyo_l]['m']     += vm
                apoyos_por_tipo[apoyo_l]['h']     += vh
                apoyos_por_tipo[apoyo_l]['total'] += contrib
            continue
        if not en_inst_l: continue
        if prog_l is None: prog_l = corr(name)

    # ── 3. BENEFICIARIOS POR MUNICIPIO ───────────────────────────────────────────
    # Los apoyos por municipio ya fueron calculados en la sección 2 desde la nueva tabla.
    # Los beneficiarios únicos por municipio (para la columna "Benef." en la tabla 3)
    # se leen desde la hoja "Beneficiarios por Municipio" con el filtro de edad/sexo.
    rows2 = list(_WB_CACHE.get('Beneficiarios por Municipio', []))

    benef_por_municipio_f = {}  # {mun_norm: {m, h, total}} — beneficiarios únicos
    inst_act2 = prog_act2 = None
    en_sdhybc2 = False
    mun_bloque2 = None

    for row in rows2:
        n = clean(row[0] if len(row) > 0 else None)
        if not n: continue
        if is_skip(n): continue

        if is_age(n):
            ak = norm_age(n)
            if ak and ak in AGE_KEYS and en_sdhybc2 and mun_bloque2:
                if mun_norm is None or _norm(mun_bloque2) == mun_norm:
                    vm = sf(row[2] if len(row) > 2 else None)
                    vh = sf(row[3] if len(row) > 3 else None)
                    vt = sf(row[4] if len(row) > 4 else None) or vm + vh
                    key = _norm(mun_bloque2)
                    if key not in benef_por_municipio_f:
                        benef_por_municipio_f[key] = {'nombre': mun_bloque2, 'm': 0.0, 'h': 0.0, 'total': 0.0}
                    if sexo == 'm':    c = vm
                    elif sexo == 'h': c = vh
                    else:             c = vt
                    benef_por_municipio_f[key]['m']     += vm
                    benef_por_municipio_f[key]['h']     += vh
                    benef_por_municipio_f[key]['total'] += c
            continue

        if is_inst(n):
            if inst_norm is None or _norm(n) == inst_norm:
                inst_act2 = n; en_sdhybc2 = True
            else:
                inst_act2 = None; en_sdhybc2 = False
            prog_act2 = None
            continue

        if is_municipio(n):
            mun_bloque2 = corr(n)
            inst_act2 = None; en_sdhybc2 = False; prog_act2 = None
            continue

        if en_sdhybc2:
            prog_act2 = corr(n)

    # ── 4. LOCALIZABLES (Tabla B hoja "Beneficiarios Localizables") ──────────
    ws_loc = 'Beneficiarios Localizables' if 'Beneficiarios Localizables' in _WB_CACHE else None
    total_loc = m_loc = h_loc = 0.0
    loc_por_rango = {r: {'m': 0.0, 'h': 0.0, 'total': 0.0} for r in (rangos_edad or list(_RANGO_COL.keys()))}

    if ws_loc:
        rows_loc = list(_WB_CACHE.get(ws_loc, []))
        INST_SET_LOC = {x.upper() for x in
                        ['DIF','ICHDII','ICHIJUV','SALUD','SDHYBC','SDHyBC','SDBYBC',
                         'SPYCI','SPyCI','CECYTECH','COESPO','COESVI','CULTURA','ICHD',
                         'ICHIMUJ','RURAL','SEECH','SEYD','SEyD','TRABAJO','TURISMO']}
        def _is_inst_l(s): return s.strip().upper() in INST_SET_LOC

        inst_act_loc = None
        ultimo_prog_loc_idx = [None]
        por_programa_loc = []

        for row in rows_loc:
            ng = clean(row[6] if len(row) > 6 else None)
            if not ng: continue
            ng_u = ng.strip().upper()
            if ng_u in ('TOTAL', 'BENEFICIARIOS LOCALIZABLES POR RANGO DE EDAD',
                        'MES CORRESPONDIENTE'): continue

            if _is_inst_l(ng):
                if inst_norm is None or _norm(ng) == inst_norm:
                    inst_act_loc = ng
                    # Fila de inst: total 0-29 (incluye sin_dato_sexo)
                    total_loc = _sum_cols(row, cols)
                    loc_inst_row_idx = rows_loc.index(row)
                else:
                    inst_act_loc = None
                continue

            if inst_act_loc is None: continue

            # Filas H/M — NOTA: en esta hoja el orden es H primero, luego M
            if ng_u in ('M', 'H'):
                v = _sum_cols(row, cols)
                if ng_u == 'M':
                    m_loc += v
                    if ultimo_prog_loc_idx[0] is not None:
                        por_programa_loc[ultimo_prog_loc_idx[0]]['m'] += v
                    for r_k, c_k in _RANGO_COL.items():
                        if c_k in cols and r_k in loc_por_rango:
                            loc_por_rango[r_k]['m'] += sf(row[c_k] if c_k < len(row) else None)
                else:
                    h_loc += v
                    if ultimo_prog_loc_idx[0] is not None:
                        por_programa_loc[ultimo_prog_loc_idx[0]]['h'] += v
                    for r_k, c_k in _RANGO_COL.items():
                        if c_k in cols and r_k in loc_por_rango:
                            loc_por_rango[r_k]['h'] += sf(row[c_k] if c_k < len(row) else None)
                continue

            # Programa
            prog_total_loc = _sum_cols(row, cols)
            por_programa_loc.append({'nombre': corr(ng), 'institucion': inst_act_loc,
                                     'm': 0.0, 'h': 0.0, 'total': prog_total_loc})
            ultimo_prog_loc_idx[0] = len(por_programa_loc) - 1

        # Completar rangos localizables
        for r_k in loc_por_rango:
            loc_por_rango[r_k]['total'] = loc_por_rango[r_k]['m'] + loc_por_rango[r_k]['h']

        if sexo == 'm':    total_loc = m_loc
        elif sexo == 'h':  total_loc = h_loc
        # else: total_loc ya fue leído de la fila de institución (incluye sn)

    # ── 5. LOCALIZABLES POR MUNICIPIO (Tabla C) ──────────────────────────────
    loc_por_municipio = {}  # {mun: {m, h, total}}
    if ws_loc:
        # Encontrar fila de encabezado de municipios (contiene "BENEFICIARIOS LOCALIZABLES POR MUNICIPIO")
        mun_col_map = {}  # {col_idx: nombre_municipio}
        hdr_c_idx = None
        rows_loc_all = rows_loc  # ya cargados arriba

        for ri, row in enumerate(rows_loc_all):
            v17 = str(row[17] if len(row) > 17 and row[17] else '').strip().upper()
            if 'BENEFICIARIOS LOCALIZABLES POR MUNICIPIO' in v17:
                hdr_c_idx = ri
                for ci in range(18, len(row)):
                    v = str(row[ci] or '').strip()
                    if v and v.upper() != 'TOTAL':
                        mun_col_map[ci] = v
                break

        if hdr_c_idx is not None:
            # Leer desde hdr_c_idx+1
            # Estructura: bloque M: INST → PROG → RANGO_EDAD
            #             luego bloque H: INST → PROG → RANGO_EDAD
            sexo_c = None
            inst_c = None
            prog_c = None

            for row in rows_loc_all[hdr_c_idx + 1:]:
                r17 = str(row[17] if len(row) > 17 and row[17] else '').strip()
                if not r17: continue
                r17u = r17.upper()

                if r17u == 'M': sexo_c = 'm'; inst_c = None; prog_c = None; continue
                if r17u == 'H': sexo_c = 'h'; inst_c = None; prog_c = None; continue
                if r17u == 'TOTAL': continue

                if r17u in {x.upper() for x in
                            ['DIF','ICHDII','ICHIJUV','SALUD','SDHYBC','SDHyBC','SDBYBC',
                             'SPYCI','SPyCI','CECYTECH','COESPO','COESVI','CULTURA','ICHD',
                             'ICHIMUJ','RURAL','SEECH','SEYD','SEyD','TRABAJO','TURISMO']}:
                    if inst_norm is None or _norm(r17) == inst_norm:
                        inst_c = r17
                    else:
                        inst_c = None
                    prog_c = None
                    continue

                if inst_c is None: continue

                # ¿Es rango de edad?
                ak = norm_age(r17)
                if ak and ak in AGE_KEYS:
                    # Leer valores por municipio
                    for ci, mun_name in mun_col_map.items():
                        if mun_norm and _norm(mun_name) != mun_norm:
                            continue
                        v = sf(row[ci] if ci < len(row) else None)
                        if v == 0: continue
                        key = mun_name
                        if key not in loc_por_municipio:
                            loc_por_municipio[key] = {'m': 0.0, 'h': 0.0, 'total': 0.0}
                        if sexo_c == 'm' and (sexo is None or sexo == 'm'):
                            loc_por_municipio[key]['m'] += v
                        elif sexo_c == 'h' and (sexo is None or sexo == 'h'):
                            loc_por_municipio[key]['h'] += v
                    continue

                # Es programa
                prog_c = corr(r17)

        # Calcular total de loc_por_municipio
        for key in loc_por_municipio:
            d = loc_por_municipio[key]
            d['total'] = d['m'] + d['h']

    # ── Construir resultado final ─────────────────────────────────────────────
    # Apoyos filtrados por sexo
    if sexo == 'm':    total_apoyos_f = m_apoyos_f
    elif sexo == 'h':  total_apoyos_f = h_apoyos_f
    else:              total_apoyos_f = m_apoyos_f + h_apoyos_f

    # Por municipio: fusionar beneficiarios (sección 3) con apoyos (nueva tabla)
    por_mun_list = []
    todos_muns = set(benef_por_municipio_f.keys()) | set(por_municipio_f.keys())
    for mun_n in todos_muns:
        bd = benef_por_municipio_f.get(mun_n, {})
        ad = por_municipio_f.get(mun_n, {})
        t  = bd.get('total', 0) or ad.get('total', 0)
        if t > 0:
            por_mun_list.append({
                'municipio':   bd.get('nombre') or ad.get('municipio') or mun_n,
                'm':           bd.get('m', 0),
                'h':           bd.get('h', 0),
                'total':       t,
                'apoyos_m':    ad.get('m', 0),
                'apoyos_h':    ad.get('h', 0),
                'apoyos_total': ad.get('total', 0),
            })
    por_mun_list.sort(key=lambda x: -x['total'])

    return {
        'total':         total_b,
        'm':             m_b,
        'h':             h_b,
        'sn':            max(0.0, total_b - m_b - h_b),
        'total_apoyos':  total_apoyos_f,
        'm_apoyos':      m_apoyos_f,
        'h_apoyos':      h_apoyos_f,
        'por_rango':     por_rango_b,
        'por_programa':  sorted([p for p in por_programa_b if p['total'] > 0 or p['m'] > 0 or p['h'] > 0],
                                 key=lambda x: -x['total']),
        'por_municipio': por_mun_list,
        'apoyos_por_programa': sorted(
            [{'nombre': k, **v} for k, v in apoyos_por_programa_nt.items() if v.get('total', 0) > 0],
            key=lambda x: -x['total']
        ),
        'localizables': {
            'total': total_loc, 'm': m_loc, 'h': h_loc,
            'por_rango': loc_por_rango,
            'por_municipio': sorted(
                [{'municipio': k, **v} for k, v in loc_por_municipio.items() if v['total'] > 0],
                key=lambda x: -x['total']
            ),
        },
        'apoyos_por_tipo': sorted(
            [{'apoyo': k, **v} for k, v in apoyos_por_tipo.items() if v['total'] > 0],
            key=lambda x: -x['total']
        ),
        'meta': {'inst_key': inst_key, 'mun_key': mun_key,
                 'sexo': sexo, 'rangos_edad': rangos_edad},
    }


instituciones, rangos, gran_total, rangos_mh_global = parse_sheet1()

# ── Salida rápida para modo --lista (menú interactivo) ────────────────────────
# Solo necesita instituciones y municipios — no ejecuta los parsers lentos.
if '--lista' in sys.argv:
    municipios_lista = parse_sheet2()
    print(json.dumps({
        'instituciones': {k: {'total': v['total'], 'm': v['m'], 'h': v['h']}
                          for k, v in instituciones.items()},
        'municipios':    [{'municipio': m['municipio'], 'total': m['total']}
                          for m in municipios_lista if not m.get('especial')],
    }, ensure_ascii=False, indent=2))
    sys.exit(0)

apoyos_g3 = parse_apoyos_g3()
apoyos_a3 = parse_apoyos_a3()
localizables          = parse_localizables()
municipios       = parse_sheet2()
apoyos, desglose = parse_sheet3_full()
indicadores      = parse_indicadores()

# Enriquecer municipios con total_apoyos desde desglose
apoyos_x_mun = {
    mun_dk: sum(a['total'] for a in aps)
    for mun_dk, aps in desglose.items()
}
for m in municipios:
    m['total_apoyos'] = apoyos_x_mun.get(_norm(m['municipio']), 0)
    mun_key = _norm(m['municipio'])   # sin tildes para cruce con hoja localizables
    m['total_localizables'] = localizables['por_municipio'].get(mun_key, 0)
    m['loc_m']              = localizables['m_por_municipio'].get(mun_key, 0)
    m['loc_h']              = localizables['h_por_municipio'].get(mun_key, 0)

# Lista simple por municipio
apm = {mun_dk: [{'apoyo':a['apoyo'],'total':a['total']} for a in aps]
       for mun_dk, aps in desglose.items()}

# Total de apoyos directo del Excel (fila TOTAL de la hoja Apoyos Otorgados)
def _leer_total_apoyos_excel():
    try:
        for row in reversed(list(_WB_CACHE.get('Apoyos Otorgados', []))):
            a = str(row[0]).strip().upper() if row[0] else ''
            if a == 'TOTAL':
                v = row[4]
                if v is not None:
                    try: return float(v)
                    except: pass
    except: pass
    return None

total_apoyos_excel = _leer_total_apoyos_excel()

# ── Filtros cruzados pre-calculados ──────────────────────────────────────────
# NOTA: Los filtros_cruzados se calculan aquí solo para generar_instituciones.py
# y generar_municipios.py. GENERAR_REPORTE.js usa --lista para el menú rápido
# y nunca llega a este bloque durante la selección interactiva.
RANGOS_0_29     = ['0-5','6-11','12-17','18-29']
RANGOS_30_64    = ['30-49','50-64']
RANGOS_MAYORES  = ['65+']

filtros_cruzados = {}
for inst_k in instituciones:
    filtros_cruzados[inst_k] = {
        'global':       calcular_filtro(instituciones, municipios, inst_key=inst_k),
        'm_total':      calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='m'),
        'h_total':      calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='h'),
        'total_0_29':   calcular_filtro(instituciones, municipios, inst_key=inst_k, rangos_edad=RANGOS_0_29),
        'total_30_64':  calcular_filtro(instituciones, municipios, inst_key=inst_k, rangos_edad=RANGOS_30_64),
        'total_65_mas': calcular_filtro(instituciones, municipios, inst_key=inst_k, rangos_edad=RANGOS_MAYORES),
        'm_0_29':       calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='m', rangos_edad=RANGOS_0_29),
        'h_0_29':       calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='h', rangos_edad=RANGOS_0_29),
        'm_30_64':      calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='m', rangos_edad=RANGOS_30_64),
        'h_30_64':      calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='h', rangos_edad=RANGOS_30_64),
        'm_65_mas':     calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='m', rangos_edad=RANGOS_MAYORES),
        'h_65_mas':     calcular_filtro(instituciones, municipios, inst_key=inst_k, sexo='h', rangos_edad=RANGOS_MAYORES),
    }

def _parse_pivot_mun_prog():
    """Lee la tabla pivot de col AJ:AK en 'Beneficiarios por Municipio'.
    Retorna {inst_norm: {prog_norm: {mun_norm: count}}}"""
    rows_piv = _WB_CACHE.get('Beneficiarios por Municipio', [])
    INST_NORM_SET = {_norm(x) for x in {
        'CECYTECH','COESPO','COESVI','DIF','ICHDII','ICHIJUV','ICHIMUJ',
        'SALUD','SDHYBC','SPYCI','CULTURA','ICHD','RURAL','SEECH','SEYD',
        'TRABAJO','TURISMO','SDHyBC','SPyCI'
    }}
    MUN_VALIDOS_NORM = {_norm(k) for k in POB_MUNICIPAL.keys()} | {_norm('NO IDENTIFICADO')}
    result = {}
    inst_act = prog_act = None
    for r in rows_piv:
        a = r[35] if len(r) > 35 else None
        b = r[36] if len(r) > 36 else None
        if a is None: continue
        na = _norm(str(a))
        if na in INST_NORM_SET:
            inst_act = na; prog_act = None
            result.setdefault(inst_act, {})
        elif inst_act and na not in INST_NORM_SET and isinstance(b, (int, float)) and na not in MUN_VALIDOS_NORM:
            prog_act = na
            result[inst_act].setdefault(prog_act, {})
        elif inst_act and prog_act and na in MUN_VALIDOS_NORM and isinstance(b, (int, float)) and b > 0:
            result[inst_act][prog_act][na] = int(b)
    return result

output = {
    'gran_total':           gran_total,
    'total_apoyos_excel':   total_apoyos_excel,
    'rangos_edad':          rangos,
    'rangos_mh_global':     rangos_mh_global,
    'instituciones':        {k:{**v} for k,v in instituciones.items()},
    'municipios':           municipios,
    'apoyos':               apoyos,
    'mun_activos':          67,
    'apoyos_por_municipio': apm,
    'desglose_municipal':   desglose,
    'apoyos_g3':            apoyos_g3,
    'apoyos_a3':            apoyos_a3,
    'pob_estatal':          POB_ESTATAL,
    'pob_municipal':        POB_MUNICIPAL,
    'localizables':         localizables,
    'indicadores':          indicadores,
    'filtros_cruzados':     filtros_cruzados,
    'pivot_mun_prog':       _parse_pivot_mun_prog(),
}
print(json.dumps(output, ensure_ascii=False, default=str, indent=2))
